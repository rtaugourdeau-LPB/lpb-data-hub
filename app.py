import re
import time
import csv
import zipfile
import json
from io import BytesIO
from typing import List, Tuple, Optional
from datetime import datetime
import math
import openpyxl
import psycopg2.extras
from datetime import date
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import requests
import psycopg2
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine, text
from rapidfuzz import process, fuzz
from unidecode import unidecode

# ============================================================
# ⚙️ CONFIG GENERALE STREAMLIT (UNE SEULE FOIS)
# ============================================================

st.set_page_config(
    page_title="Outils Data LPB",
    page_icon="🧱",
    layout="wide",
)

# Date du jour pour les exports
TODAY_STR = datetime.now().strftime("%Y-%m-%d")
MAX_ROWS_EXPORT = 1_000_000      # limite globale (pour CSV/ZIP et Excel)
ROWS_PER_CSV_FILE = 200_000      # lignes max par fichier CSV dans le ZIP
MAX_ROWS_EXCEL = 200_000         # limite stricte pour Excel (évite les xlsx énormes)

# ============================================================
# 🔐 CONFIG VIA st.secrets (Streamlit)
# ============================================================

# --- Airtable (Votes) ---
AIRTABLE_TOKEN   = st.secrets["AIRTABLE_TOKEN"]
AIRTABLE_BASE_ID = st.secrets["AIRTABLE_BASE_ID"]
AIR_H = {"Authorization": f"Bearer {AIRTABLE_TOKEN}"}

# --- Notion (Data Hub) ---
NOTION_TOKEN     = st.secrets["NOTION_TOKEN"]
DB_BAC_SABLE     = st.secrets["DB_BAC_SABLE"]
DB_NOTION_PROJET = st.secrets["DB_NOTION_PROJET"]

HEADERS = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Notion-Version": "2022-06-28",
    "Content-Type": "application/json",
}

# --- PostgreSQL (BO / Data Hub) ---
PG_HOST     = st.secrets["PG_HOST"]
PG_PORT     = st.secrets["PG_PORT"]
PG_DB       = st.secrets["PG_DB"]
PG_USER     = st.secrets["PG_USER"]
PG_PASSWORD = st.secrets["PG_PASSWORD"]
PG_SSLMODE  = st.secrets.get("PG_SSLMODE", "require")

# URL publique projet LPB (utilisée dans page_votes)
LPB_PROJECT_URL = st.secrets.get(
    "LPB_PROJECT_URL",
    "https://app.lapremierebrique.fr/fr/projects/{project_id}",
)

# variables utilisées par psycopg2
PGHOST     = PG_HOST
PGPORT     = PG_PORT
PGDATABASE = PG_DB
PGUSER     = PG_USER
PGPASSWORD = PG_PASSWORD
PGSSLMODE  = PG_SSLMODE

#========
# 🧠 ÉTAT GLOBAL STREAMLIT (pour Data Hub)
# =============================================================================

# API Notion
if "api_df" not in st.session_state:
    st.session_state.api_df = None

# Import ZIP/CSV
if "imp_df" not in st.session_state:
    st.session_state.imp_df = None
if "imp_filename" not in st.session_state:
    st.session_state.imp_filename = None

# Back-office (PostgreSQL)
if "bo_df" not in st.session_state:
    st.session_state.bo_df = None
if "bo_table_id" not in st.session_state:
    st.session_state.bo_table_id = None  # full_name = schema.table
if "bo_tables_df" not in st.session_state:
    st.session_state.bo_tables_df = None


# =============================================================================
# 🔧 FONCTIONS COMMUNES (utilisées par les 2 modules)
# =============================================================================

def ensure_ok(r: requests.Response):
    """Raise si HTTP != 2xx, avec détail."""
    try:
        r.raise_for_status()
    except Exception as e:
        try:
            detail = r.json()
        except Exception:
            detail = r.text
        raise RuntimeError(f"HTTP {r.status_code}: {detail}") from e


# ======================= Utils strings (Votes) =======================

def normalize_name(s: str) -> str:
    """Normalisation douce pour comparer les noms."""
    if s is None:
        return ""
    s2 = unidecode(str(s)).lower().strip()
    s2 = re.sub(r"\s+", " ", s2)
    s2 = re.sub(r"^(le|la|les|l')\s+", "", s2)
    return s2


def normalize_email(s: str) -> str:
    if s is None:
        return ""
    return unidecode(str(s)).strip().lower()


def extract_dept(token: str) -> Optional[int]:
    """Extrait (DD) ou (DDD) en fin de libellé → int, sinon None."""
    m = re.search(r"\((\d{2,3})\)\s*$", str(token).strip())
    return int(m.group(1)) if m else None


def make_nom_airtable(name, department) -> str:
    """Construit 'Nom (DD)' à partir de BO (name, department)."""
    n = "" if pd.isna(name) else str(name).strip()
    if pd.isna(department) or str(department).strip() == "":
        return n
    try:
        dep = int(department)
    except Exception:
        dep = str(department).strip()
    return f"{n} ({dep})"


# ======================= Patterns globaux prolongation / pouvoir =======================

PROLONG_PATTERN = re.compile(
    r"(prolongation|pr[ée]pa|pr[ée]par|pr[ée]paration|prepa|ne\s*se\s*pron|nspp)",
    re.I,
)
POUVOIR_PATTERN = re.compile(
    r"(pouvoir|procuration|proxy)",
    re.I,
)

# =============================================================================
# 🔧 FONCTIONS AIRTABLE (Votes)
# =============================================================================

def list_tables_with_views(base_id: str) -> List[dict]:
    """Retourne la métadonnée Airtable de toutes les tables de la base."""
    if not AIR_H:
        return []
    r = requests.get(
        f"https://api.airtable.com/v0/meta/bases/{base_id}/tables",
        headers=AIR_H,
        timeout=30,
    )
    ensure_ok(r)
    return r.json().get("tables", []) or []


def table_has_prolongation_or_pouvoir(table_meta: dict) -> bool:
    """
    Ne garde que les tables qui ont AU MOINS une colonne liée à 'prolongation' OU 'pouvoir'.
    """
    for f in table_meta.get("fields", []) or []:
        name = str(f.get("name", ""))
        if PROLONG_PATTERN.search(name) or POUVOIR_PATTERN.search(name):
            return True
    return False


def fetch_view_records(
    base_id: str,
    table_id_or_name: str,
    view_id_or_name: str,
    page_size: int = 100,
) -> List[dict]:
    if not AIR_H:
        return []
    url = f"https://api.airtable.com/v0/{base_id}/{table_id_or_name}"
    params = {"pageSize": page_size, "view": view_id_or_name}
    out, offset = [], None
    while True:
        if offset:
            params["offset"] = offset
        r = requests.get(url, headers=AIR_H, params=params, timeout=60)
        ensure_ok(r)
        data = r.json()
        out.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            break
        time.sleep(0.12)  # respect soft-rate
    return out


def flatten(records: List[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    rows = []
    for r in records:
        d = {"_air_id": r.get("id"), "_air_createdTime": r.get("createdTime")}
        d.update(r.get("fields", {}) or {})
        rows.append(d)
    return pd.DataFrame(rows)


# =============================================================================
# 🔧 FONCTIONS BO (Votes) — SQLAlchemy
# =============================================================================

@st.cache_resource(show_spinner=False)
def get_engine():
    """Engine SQLAlchemy sur le BO (read-replica)."""
    if not all([PG_HOST, PG_DB, PG_USER, PG_PASSWORD]):
        return None
    uri = (
        f"postgresql+psycopg2://{PG_USER}:{PG_PASSWORD}"
        f"@{PG_HOST}:{PG_PORT}/{PG_DB}?sslmode={PG_SSLMODE}"
    )
    return create_engine(uri, pool_pre_ping=True)


@st.cache_data(show_spinner=True)
def load_projects_df() -> pd.DataFrame:
    eng = get_engine()
    if eng is None:
        return pd.DataFrame(
            columns=["id", "name", "department", "nom_airtable", "name_norm", "nom_airtable_norm"]
        )
    q = text(
        """
        SELECT id, name, department
        FROM public.projects
        WHERE name IS NOT NULL
        """
    )
    df = pd.read_sql(q, eng)
    df["nom_airtable"] = df.apply(
        lambda r: make_nom_airtable(r["name"], r["department"]),
        axis=1,
    )
    df["name_norm"] = df["name"].apply(normalize_name)
    df["nom_airtable_norm"] = df["nom_airtable"].apply(normalize_name)
    return df


@st.cache_data(show_spinner=True)
def load_subs_for_project(project_id: int) -> pd.DataFrame:
    eng = get_engine()
    if eng is None:
        return pd.DataFrame(
            columns=[
                "subscription_id",
                "users_profile_id",
                "project_id",
                "subscribed_at",
                "email_normalized",
                "email_raw",
            ]
        )
    q = text(
        """
        SELECT
            s.id AS subscription_id,
            s.users_profile_id AS users_profile_id,
            s.project_id,
            s.created_at AS subscribed_at,
            lower(trim(u.email)) AS email_normalized,
            u.email AS email_raw
        FROM public.subscriptions s
        JOIN public.users_profiles up ON up.id = s.users_profile_id
        LEFT JOIN public.users u ON u.id = up.user_id
        WHERE s.status <> 'canceled'
          AND s.project_id = :pid
          AND u.email IS NOT NULL
        """
    )
    return pd.read_sql(q, eng, params={"pid": int(project_id)})

@st.cache_data(show_spinner=True)
def load_invest_amounts_for_project(project_id: int) -> pd.DataFrame:
    """
    Montant investi total par email sur ce projet (pondération).
    IMPORTANT : adapte le champ montant si besoin (amount, amount_cents, etc.)
    """
    eng = get_engine()
    if eng is None:
        return pd.DataFrame(columns=["email_normalized", "invest_amount_eur"])

    q = text(
        """
        SELECT
            lower(trim(u.email)) AS email_normalized,
            SUM(s.amount) AS invest_amount_eur
        FROM public.subscriptions s
        JOIN public.users_profiles up ON up.id = s.users_profile_id
        LEFT JOIN public.users u ON u.id = up.user_id
        WHERE s.status <> 'canceled'
          AND s.project_id = :pid
          AND u.email IS NOT NULL
        GROUP BY 1
        """
    )
    return pd.read_sql(q, eng, params={"pid": int(project_id)})
    
# =============================================================================
# 🔧 Matching Projet (Airtable → BO)
# =============================================================================

def guess_candidates_from_label(
    label: str,
    projects_df: pd.DataFrame,
    topn: int = 10,
) -> List[Tuple[int, str, float]]:
    cand: List[Tuple[int, str, float]] = []
    label = (label or "").strip()
    dept = extract_dept(label)
    label_no_dep = re.sub(r"\s*\(\d{2,3}\)\s*$", "", label).strip()

    # 1) exact name + dept
    if dept is not None:
        m1 = projects_df.query("department == @dept and name == @label_no_dep")
        for _, r in m1.iterrows():
            cand.append((int(r["id"]), r["nom_airtable"], 100.0))

    # 2) exact name seul
    m2 = projects_df.query("name == @label_no_dep")
    for _, r in m2.iterrows():
        tup = (int(r["id"]), r["nom_airtable"], 99.0)
        if tup not in cand:
            cand.append(tup)

    # 3) exact normalisé
    label_norm = normalize_name(label_no_dep)
    m3 = projects_df[projects_df["name_norm"] == label_norm]
    for _, r in m3.iterrows():
        tup = (int(r["id"]), r["nom_airtable"], 98.0)
        if tup not in cand:
            cand.append(tup)

    # 4) fuzzy sur nom_airtable puis name
    choices = projects_df["nom_airtable"].tolist()
    for (disp, score, pos) in process.extract(
        label,
        choices,
        scorer=fuzz.token_set_ratio,
        limit=topn,
    ):
        r = projects_df.iloc[pos]
        tup = (int(r["id"]), r["nom_airtable"], float(score))
        if tup not in cand:
            cand.append(tup)

    choices2 = projects_df["name"].tolist()
    for (disp, score, pos) in process.extract(
        label_no_dep,
        choices2,
        scorer=fuzz.token_set_ratio,
        limit=topn,
    ):
        r = projects_df.iloc[pos]
        tup = (int(r["id"]), r["nom_airtable"], float(score) - 0.5)
        if tup not in cand:
            cand.append(tup)

    return sorted(cand, key=lambda x: x[2], reverse=True)[:topn]


def pick_project_id_from_airtable(
    df_view: pd.DataFrame,
    projects_df: pd.DataFrame,
    view_name: str,
    table_name: Optional[str] = None,
) -> Tuple[int, str]:
    """
    Ordre de résolution :
    1) colonne 'nom de projet' dans la vue (Nom du projet, Projet, Project name…)
    2) colonne URL → extrait /projects/<id>
    3) fallback depuis nom de vue ou de table
    """

    candidate_label = None

    # (1) colonne ‘nom projet’
    name_cols = [
        c
        for c in df_view.columns
        if re.search(r"(nom.*projet|projet.*nom|project.?name)", str(c), re.I)
    ]
    if name_cols:
        s = df_view[name_cols[0]].dropna().astype(str).str.strip()
        if not s.empty:
            candidate_label = s.value_counts().idxmax()

    # (2) colonne lien → ID
    if candidate_label is None:
        url_cols = [
            c for c in df_view.columns if re.search(r"(url|lien|link)", str(c), re.I)
        ]
        if url_cols:
            s = df_view[url_cols[0]].dropna().astype(str)
            m = s.str.extract(r"/projects/(\d+)", expand=False).dropna()
            if not m.empty:
                pid = int(m.iloc[0])
                disp = projects_df.loc[projects_df["id"] == pid, "nom_airtable"]
                if not disp.empty:
                    return pid, disp.iloc[0]

    # (3) fallbacks intelligents
    def extract_from_title(txt: str) -> str:
        if not txt:
            return ""
        m = re.search(r"Projet\s+(.+)$", txt, flags=re.I)
        return m.group(1).strip() if m else txt.strip()

    if candidate_label is None:
        looks_like_view_id = bool(re.match(r"^Vue\s+viw[A-Za-z0-9]+$", str(view_name)))
        if table_name and (looks_like_view_id or "Projet" in str(table_name)):
            candidate_label = extract_from_title(str(table_name))
        else:
            candidate_label = extract_from_title(str(view_name))

    st.write("Libellé de croisement Airtable :", f"{candidate_label}")

    cands = guess_candidates_from_label(candidate_label, projects_df, topn=5)

    # Dédupliquer par project_id en gardant le meilleur score
    best = {}
    for pid, disp, score in cands:
        if (pid not in best) or (score > best[pid][1]):
            best[pid] = (disp, float(score))
    cands_unique = [(pid, disp, score) for pid, (disp, score) in best.items()]

    # ⚖️ Tri : d'abord les projets NON "Canceled", puis score décroissant, puis libellé
    def is_canceled(name: str) -> int:
        # 0 = projet normal, 1 = projet "Canceled"
        return 1 if re.search(r"\bcanceled\b", str(name), flags=re.I) else 0

    cands_unique.sort(
        key=lambda x: (
            is_canceled(x[1]),  # non-canceled (0) avant canceled (1)
            -x[2],              # score décroissant
            x[1],               # puis libellé
        )
    )

    options = [
        f"{pid} — {disp} (score {score:.1f})"
        for pid, disp, score in cands_unique
    ]
    options.append("Saisir manuellement")

    choice = st.selectbox(
        "Confirme le projet exact BO (les plus probables) :",
        options,
        index=0,
    )

    if choice == "Saisir manuellement":
        pid = st.number_input("ID projet LPB :", min_value=1, step=1)
        if pid in projects_df["id"].values:
            disp = projects_df.loc[projects_df["id"] == pid, "nom_airtable"].iloc[0]
        else:
            disp = f"Projet {int(pid)}"
        return int(pid), disp

    pid = int(re.match(r"^(\d+)\s—", choice).group(1))
    disp = projects_df.loc[projects_df["id"] == pid, "nom_airtable"].iloc[0]
    return pid, disp


# =============================================================================
# 🔧 Détection colonnes (emails / prolongation / pouvoir) — Votes
# =============================================================================

def detect_email_columns(df: pd.DataFrame) -> List[str]:
    cols = [c for c in df.columns if re.search(r"mail|e-?mail", str(c), re.I)]
    if cols:
        return cols
    # fallback: première colonne contenant '@'
    for c in df.columns:
        try:
            if df[c].astype(str).str.contains("@").any():
                return [c]
        except Exception:
            pass
    return []


def detect_prolongation_column(df: pd.DataFrame) -> Optional[str]:
    for c in df.columns:
        if PROLONG_PATTERN.search(str(c)):
            return c
    return None


def detect_pouvoir_column(df: pd.DataFrame) -> Optional[str]:
    for c in df.columns:
        if POUVOIR_PATTERN.search(str(c)):
            return c
    return None


def standardize_prolongation(val) -> str:
    """
    Prolongation = réponse à :
    - si colonne 'pouvoir' présente :
        « Êtes-vous d'accord pour accorder la prolongation ? »
    - sinon :
        « Êtes-vous d'accord pour appliquer la prolongation, avec application des 5% de pénalité ? »
    """
    if pd.isna(val):
        return "Non renseigné"
    s = unidecode(str(val)).strip().lower()
    if s in {"oui", "o", "yes", "y", "true", "1"}:
        return "Oui"
    if s in {"non", "n", "no", "false", "0"}:
        return "Non"
    if re.search(r"ne\s*se\s*pron", s) or s in {"nspp", "ne se prononce pas"}:
        return "Ne se prononce pas"
    return str(val)


def standardize_pouvoir(val) -> str:
    """
    Pouvoir = réponse à : « Êtes-vous d'accord pour NE PAS APPLIQUER les pénalités ? »
    """
    if pd.isna(val):
        return "Non renseigné"
    s = unidecode(str(val)).strip().lower()
    if s in {"oui", "o", "yes", "y", "true", "1"}:
        return "Oui"
    if s in {"non", "n", "no", "false", "0"}:
        return "Non"
    if re.search(r"ne\s*se\s*pron", s) or s in {"nspp", "ne se prononce pas"}:
        return "Ne se prononce pas"
    if s in {"", "nan"}:
        return "Non renseigné"
    return str(val)


def build_votes_email_flags(
    df_view: pd.DataFrame,
    email_cols: List[str],
    prolong_col: Optional[str],
    pouvoir_col: Optional[str],
) -> pd.DataFrame:
    """
    Agrégé par e-mail :
    - email_normalized
    - email_raw_example
    - prolongation (mode standardisé)
    - pouvoir (mode standardisé)
    - n_occur
    """
    if not email_cols:
        return pd.DataFrame(
            columns=["email_normalized", "email_raw_example", "prolongation", "pouvoir"]
        )

    melted = []
    for c in email_cols:
        if c not in df_view.columns:
            continue
        s = df_view[[c]].copy()
        s = s.dropna()
        if s.empty:
            continue
        s["email_raw"] = s[c].astype(str)
        s["email_normalized"] = s["email_raw"].map(normalize_email)
        s["source_col"] = c

        if prolong_col and prolong_col in df_view.columns:
            s["prolongation_raw"] = df_view.loc[s.index, prolong_col]
        else:
            s["prolongation_raw"] = None

        if pouvoir_col and pouvoir_col in df_view.columns:
            s["pouvoir_raw"] = df_view.loc[s.index, pouvoir_col]
        else:
            s["pouvoir_raw"] = None

        melted.append(
            s[
                [
                    "email_raw",
                    "email_normalized",
                    "source_col",
                    "prolongation_raw",
                    "pouvoir_raw",
                ]
            ]
        )

    if not melted:
        return pd.DataFrame(
            columns=["email_normalized", "email_raw_example", "prolongation", "pouvoir"]
        )

    tmp = pd.concat(melted, ignore_index=True)
    tmp = tmp[tmp["email_normalized"].str.contains("@", na=False)]

    tmp["prolongation_std"] = tmp["prolongation_raw"].apply(standardize_prolongation)
    tmp["pouvoir_std"] = tmp["pouvoir_raw"].apply(standardize_pouvoir)

    agg = (
        tmp.groupby("email_normalized", as_index=False)
        .agg(
            n_occur=("email_normalized", "size"),
            email_raw_example=("email_raw", "first"),
            prolongation=(
                "prolongation_std",
                lambda x: x.dropna().mode().iloc[0]
                if not x.dropna().empty
                else "Non renseigné",
            ),
            pouvoir=(
                "pouvoir_std",
                lambda x: x.dropna().mode().iloc[0]
                if not x.dropna().empty
                else "Non renseigné",
            ),
        )
        .sort_values("email_normalized")
    )
    return agg


# =============================================================================
# 🔧 FONCTIONS EXPORT (Data Hub)
# =============================================================================

def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data"):
    """
    Export propre vers Excel :
    - Table structurée avec style
    - Largeur de colonnes auto
    - Ligne d’en-tête figée
    - Filtres auto
    - Hyperliens désactivés
    """
    out = BytesIO()
    df_export = df.copy()

    # Colonnes plutôt texte
    text_col_keywords = (
        "url",
        "link",
        "href",
        "file",
        "image",
        "img",
        "path",
        "uuid",
        "token",
        "hash",
        "id",
        "code",
        "ref",
        "reference",
        "email",
        "mail",
        "phone",
        "tel",
    )

    for col in df_export.columns:
        s = df_export[col]
        if not (pd.api.types.is_object_dtype(s) or pd.api.types.is_string_dtype(s)):
            continue

        col_lower = str(col).lower()
        force_text = any(k in col_lower for k in text_col_keywords)

        s_str = s.astype(str)

        url_like = s_str.str.startswith(("http://", "https://", "www."))
        email_like = s_str.str.contains("@", na=False)
        long_like = s_str.str.len() > 120

        if url_like.mean() > 0.1 or email_like.mean() > 0.3 or long_like.mean() > 0.3:
            force_text = True

        def to_excel_safe(v):
            # ⛔️ ne pas utiliser pd.isna ici (problème avec les arrays)
            if v is None or (isinstance(v, float) and math.isnan(v)):
                return None
            sv = str(v)
            if sv.startswith(("=", "+", "-", "@")):
                sv = "'" + sv
            if sv.startswith(("http://", "https://", "www.")):
                sv = "'" + sv
            return sv

        if force_text:
            df_export[col] = s_str.map(to_excel_safe)
        else:
            def sanitize(v):
                # ⛔️ idem : uniquement scalaires
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    return None
                sv = v
                if isinstance(sv, str):
                    if sv.startswith(("=", "+", "-", "@")):
                        sv = "'" + sv
                    if sv.startswith(("http://", "https://", "www.")) and len(sv) > 255:
                        sv = "'" + sv
                return sv

            df_export[col] = s.map(sanitize)

    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        n_rows, n_cols = df_export.shape

        # Largeur de colonnes auto
        for col_idx, col_name in enumerate(df_export.columns):
            col_series = df_export[col_name].astype(str).fillna("")
            sample = col_series.head(500)
            max_len_value = sample.map(len).max() if not sample.empty else 0
            max_len = max(len(str(col_name)), max_len_value)
            base_width = min(max_len + 2, 60)
            width = base_width + 4
            worksheet.set_column(col_idx, col_idx, width)

        worksheet.freeze_panes(1, 0)

        if n_cols > 0:
            table_columns = [{"header": str(col)} for col in df_export.columns]
            worksheet.add_table(
                0,
                0,
                n_rows,
                n_cols - 1,
                {
                    "columns": table_columns,
                    "style": "Table Style Medium 2",
                    "autofilter": True,
                },
            )

        worksheet.set_zoom(100)

    return out.getvalue()


def df_to_csv_bytes(df: pd.DataFrame):
    """Export CSV simple, compatible Excel FR."""
    csv_str = df.to_csv(index=False, sep=";", encoding="utf-8-sig")
    return csv_str.encode("utf-8-sig")

def chunk_df_to_zip_csv_bytes(df: pd.DataFrame, rows_per_file: int, base_name: str) -> bytes:
    """Découpe un DF en plusieurs CSV et renvoie un ZIP (bytes)."""
    out = BytesIO()
    with zipfile.ZipFile(out, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        n = len(df)
        n_files = max(1, math.ceil(n / rows_per_file))
        for i in range(n_files):
            start = i * rows_per_file
            end = min((i + 1) * rows_per_file, n)
            chunk = df.iloc[start:end]
            csv_bytes = df_to_csv_bytes(chunk)
            zf.writestr(f"{base_name}_{i+1:03d}.csv", csv_bytes)
    return out.getvalue()

# =============================================================================
# 🔧 FONCTIONS NOTION (Data Hub)
# =============================================================================
######################################################################################
@st.cache_data(show_spinner=False)
def load_notion_projects_df() -> pd.DataFrame:
    """
    Charge la base Notion des projets (DB_NOTION_PROJET) et ajoute une colonne
    ID_BO_clean (entier) basée sur 'ID Back-office' ou '🔑 ID Back-office'.
    """
    try:
        schema_order = get_database_schema(DB_NOTION_PROJET)
        results = get_all_rows(DB_NOTION_PROJET)
    except Exception as e:
        st.warning(f"Impossible de charger la base Notion projets : {e}")
        return pd.DataFrame()

    if not results:
        return pd.DataFrame()

    df = notion_to_df(results, schema_order)

    # Cherche le bon nom de colonne pour l'ID BO
    id_bo_col = None
    for col in df.columns:
        if col in ("ID Back-office", "🔑 ID Back-office"):
            id_bo_col = col
            break

    if id_bo_col is not None:
        df["ID_BO_clean"] = (
            pd.to_numeric(df[id_bo_col], errors="coerce")
              .astype("Int64")  # entier nullable
        )
    else:
        # Colonne absente : on met juste une colonne vide
        df["ID_BO_clean"] = pd.Series([pd.NA] * len(df), dtype="Int64")

    return df

from datetime import datetime

def get_vote_result_date_for_project(project_id: int) -> Optional[str]:
    df = load_notion_projects_df()
    if df is None or df.empty:
        return None

    try:
        pid = int(project_id)
    except Exception:
        return None

    match = df[df["ID_BO_clean"] == pid]
    if match.empty:
        return None

    row = match.iloc[0]
    val = row.get("Date résultat de vote")

    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None

    raw = str(val).strip()

    # cas "start → end" (plage de dates)
    if "→" in raw:
        start_raw, end_raw = [x.strip() for x in raw.split("→", 1)]
        def fmt(d):
            try:
                return datetime.fromisoformat(d.split("T")[0]).strftime("%d/%m/%Y")
            except Exception:
                return d
        return f"{fmt(start_raw)} → {fmt(end_raw)}"

    # cas simple : une seule date
    try:
        d = datetime.fromisoformat(raw.split("T")[0])
        return d.strftime("%d/%m/%Y")
    except Exception:
        return raw
######################################################################################

def get_database_schema(db_id: str):
    """Récupère la liste (nom, type) de toutes les propriétés de la base, dans l'ordre API."""
    url = f"https://api.notion.com/v1/databases/{db_id}"
    r = requests.get(url, headers=HEADERS)
    if r.status_code != 200:
        raise Exception(f"Erreur schéma Notion : {r.text}")
    data = r.json()
    props = data.get("properties", {})
    return [(name, definition.get("type")) for name, definition in props.items()]


def get_all_rows(db_id: str):
    """Récupère toutes les pages d'une base Notion (pagination)."""
    url = f"https://api.notion.com/v1/databases/{db_id}/query"
    payload = {"page_size": 100}
    results = []
    while True:
        r = requests.post(url, headers=HEADERS, json=payload)
        if r.status_code != 200:
            raise Exception(f"Erreur Notion : {r.text}")
        data = r.json()
        results.extend(data.get("results", []))
        if not data.get("has_more"):
            break
        payload["start_cursor"] = data.get("next_cursor")
    return results


def _join_plain_text(rich_list):
    if not rich_list:
        return None
    try:
        return "".join([r.get("plain_text", "") for r in rich_list]) or None
    except Exception:
        return None


def parse_property_value(prop: dict):
    """
    Convertit un property Notion -> valeur simple.
    On couvre au max les types, le reste en JSON pour ne rien perdre.
    """
    if not isinstance(prop, dict):
        return prop
    t = prop.get("type")
    value = prop.get(t)

    # texte
    if t in ("title", "rich_text"):
        return _join_plain_text(value)

    # select / status
    if t in ("select", "status"):
        return value["name"] if value else None

    # multi_select
    if t == "multi_select":
        if not value:
            return None
        return ", ".join(v.get("name", "") for v in value)

    # number
    if t == "number":
        return value

    # date
    if t == "date":
        if not value:
            return None
        start = value.get("start")
        end = value.get("end")
        return f"{start} → {end}" if end else start

    # checkbox
    if t == "checkbox":
        return value

    # url / email / phone
    if t in ("url", "email", "phone_number"):
        return value

    # people
    if t == "people":
        if not value:
            return None
        names = [
            p.get("name") or p.get("id") for p in value if p.get("name") or p.get("id")
        ]
        return ", ".join(names) if names else None

    # files
    if t == "files":
        if not value:
            return None
        parts = []
        for f in value:
            name = f.get("name")
            f_type = f.get("type")
            url = None
            if f_type == "file":
                url = f.get("file", {}).get("url")
            elif f_type == "external":
                url = f.get("external", {}).get("url")
            if name and url:
                parts.append(f"{name} ({url})")
            elif url:
                parts.append(url)
            elif name:
                parts.append(name)
        return ", ".join(parts) if parts else None

    # formula
    if t == "formula":
        if not value:
            return None
        f_type = value.get("type")
        if f_type in ("string", "number", "boolean"):
            return value.get(f_type)
        if f_type == "date":
            d = value.get("date")
            if not d:
                return None
            s = d.get("start")
            e = d.get("end")
            return f"{s} → {e}" if e else s
        return json.dumps(value, ensure_ascii=False)

    # relation
    if t == "relation":
        if not value:
            return None
        return ", ".join(rel.get("id", "") for rel in value) or None

    # rollup
    if t == "rollup":
        if not value:
            return None
        r_type = value.get("type")
        if r_type in ("number", "boolean"):
            return value.get(r_type)
        if r_type == "date":
            d = value.get("date")
            if not d:
                return None
            s = d.get("start")
            e = d.get("end")
            return f"{s} → {e}" if e else s
        if r_type == "array":
            items_txt = []
            for it in value.get("array", []):
                it_type = it.get("type")
                it_val = it.get(it_type)
                if isinstance(it_val, list):
                    txt = _join_plain_text(it_val)
                    if txt:
                        items_txt.append(txt)
                elif isinstance(it_val, dict):
                    if "name" in it_val:
                        items_txt.append(it_val["name"])
                else:
                    if it_val is not None:
                        items_txt.append(str(it_val))
            return ", ".join(items_txt) if items_txt else None
        return json.dumps(value, ensure_ascii=False)

    # created_time / last_edited_time
    if t in ("created_time", "last_edited_time"):
        return value

    # created_by / last_edited_by
    if t in ("created_by", "last_edited_by"):
        if not value:
            return None
        return value.get("name") or value.get("id")

    # unique_id
    if t == "unique_id":
        if not value:
            return None
        prefix = value.get("prefix") or ""
        number = value.get("number")
        if prefix and number is not None:
            return f"{prefix}-{number}"
        if number is not None:
            return str(number)
        return json.dumps(value, ensure_ascii=False)

    # fallback : JSON
    return json.dumps({k: v for k, v in prop.items() if k != "id"}, ensure_ascii=False)


def notion_to_df(results, schema_order):
    """Construit un DataFrame API Notion avec une colonne par propriété."""
    prop_names = [name for name, _ in schema_order]
    rows = []
    for page in results:
        props = page.get("properties", {})
        row = {}
        for name in prop_names:
            prop_value = props.get(name)
            row[name] = (
                parse_property_value(prop_value) if prop_value is not None else None
            )
        rows.append(row)
    return pd.DataFrame(rows, columns=prop_names)


# =============================================================================
# 🔧 FONCTIONS IMPORT ZIP / CSV (Data Hub)
# =============================================================================

def extract_csv_recursive(zip_file):
    """
    Ouvre un ZIP Notion (ZIP dans ZIP possible) et renvoie une liste
    de (nom, DataFrame) pour tous les CSV trouvés.
    """
    csv_files = []

    def explore(zip_bytes):
        z = zipfile.ZipFile(zip_bytes)
        for name in z.namelist():
            if name.lower().endswith(".csv"):
                with z.open(name) as f:
                    csv_files.append((name, pd.read_csv(f)))
            elif name.lower().endswith(".zip"):
                nested_bytes = BytesIO(z.read(name))
                explore(nested_bytes)

    explore(zip_file)
    return csv_files


# =============================================================================
# 🔧 FONCTIONS POSTGRESQL (Data Hub) — psycopg2
# =============================================================================

@st.cache_resource(show_spinner=False)
def get_pg_connection():
    conn = psycopg2.connect(
        host=PGHOST,
        port=PGPORT,
        dbname=PGDATABASE,
        user=PGUSER,
        password=PGPASSWORD,
        sslmode=PGSSLMODE,
    )
    return conn


@st.cache_data(show_spinner=False)
def list_pg_tables():
    """Renvoie un DataFrame (table_schema, table_name) pour toutes les tables utilisateur."""
    conn = get_pg_connection()
    query = """
    SELECT table_schema, table_name
    FROM information_schema.tables
    WHERE table_type = 'BASE TABLE'
      AND table_schema NOT IN ('pg_catalog', 'information_schema')
    ORDER BY table_schema, table_name;
    """
    return pd.read_sql(query, conn)


@st.cache_data(show_spinner=False)
def read_pg_table(schema: str, table: str) -> pd.DataFrame:
    """Lit la table PostgreSQL sans planter sur les dates invalides (date/timestamp => texte)."""
    conn = get_pg_connection()

    # 1) récupérer les colonnes + types
    meta_q = """
    SELECT column_name, data_type
    FROM information_schema.columns
    WHERE table_schema = %s
      AND table_name = %s
    ORDER BY ordinal_position;
    """
    meta = pd.read_sql(meta_q, conn, params=(schema, table))

    # 2) construire un SELECT "safe"
    safe_cols = []
    for _, r in meta.iterrows():
        col = r["column_name"]
        typ = str(r["data_type"]).lower()
        if typ in ("date", "timestamp without time zone", "timestamp with time zone"):
            safe_cols.append(f'"{col}"::text AS "{col}"')
        else:
            safe_cols.append(f'"{col}"')

    query = f'SELECT {", ".join(safe_cols)} FROM "{schema}"."{table}"'
    return pd.read_sql(query, conn)


# =============================================================================
# 📄 MODULE 1 : PAGE "Votes Airtable ↔ Souscriptions BO"
# =============================================================================

def page_votes():
    st.title("📊 Vérification des votes Airtable")

    # ----- Sidebar spécifique Votes -----
    with st.sidebar:
        st.header("⚙️ Votes Airtable ↔ BO")
        st.caption(
            "Lien de vérification Airtable : "
            "https://airtable.com/appjOQoptI7Av1obe/tblpoKvFoobl4yej0/viwbPANJvZO7AVX3A?blocks=hide"
        )

    with st.spinner("Chargement référentiel projets BO…"):
        projects_df = load_projects_df()

    # 1) Lister les vues Airtable contenant 'vote' ET dont la TABLE a prolongation ou pouvoir
    if not AIRTABLE_BASE_ID or not AIR_H:
        st.error(
            "Configure AIRTABLE_TOKEN / AIRTABLE_BASE_ID dans le fichier pour interroger Airtable."
        )
        st.stop()

    tables = list_tables_with_views(AIRTABLE_BASE_ID)
    vote_re = re.compile(r"\bvote[s]?\b", re.I)
    views = []
    for t in tables:
        if not table_has_prolongation_or_pouvoir(t):
            continue
        for v in t.get("views", []) or []:
            if vote_re.search(v.get("name", "")):
                views.append((t["name"], t["id"], v["name"], v["id"]))
    views = sorted(views, key=lambda x: (x[0].lower(), x[2].lower()))

    st.subheader("1) Sélection de la vue Airtable (tableau de votes)")

    if not views:
        st.warning(
            "Aucune vue éligible trouvée (nom contenant 'vote' ET table avec colonne "
            "'prolongation' ou 'pouvoir')."
        )
        st.stop()

    labels = [f"{tname} • {vname}" for (tname, tid, vname, vid) in views]

    q = st.text_input(
        "Rechercher par mot-clé ou coller l’URL d’une vue Airtable (optionnel)",
        placeholder="ex: Solaire ou https://airtable.com/app.../tbl.../viw...",
    ).strip()

    selected_by_url = None
    if q.startswith("http"):
        m = re.search(r"airtable\.com/([A-Za-z0-9]+)/([A-Za-z0-9]+)/([A-Za-z0-9]+)", q)
        if m:
            base_from_url, tbl_from_url, view_from_url = m.groups()
            if base_from_url == AIRTABLE_BASE_ID:
                selected_by_url = (tbl_from_url, view_from_url)
            else:
                st.info(
                    "⚠️ L’URL ne correspond pas à la base configurée dans ce script "
                    "(AIRTABLE_BASE_ID)."
                )

    # --- Sélection de la vue ---
    if selected_by_url:
        tid, vid = selected_by_url
        tname = next((t for (t, _tid, _v, _vid) in views if _tid == tid), f"Table {tid}")
        vname = next((v for (_t, _tid, v, _vid) in views if _vid == vid), f"Vue {vid}")
    else:
        if q and not q.startswith("http"):
            filt_idx = [i for i, lab in enumerate(labels) if q.lower() in lab.lower()]
            if not filt_idx:
                st.info("Aucun résultat pour ce filtre. Affichage de toutes les vues éligibles.")
                options = labels
                idx_map = list(range(len(labels)))
            else:
                options = [labels[i] for i in filt_idx]
                idx_map = filt_idx
        else:
            options = labels
            idx_map = list(range(len(labels)))

        choice = st.selectbox(
            "Choisis un tableau de votes (vue Airtable) :",
            options,
            index=0,
        )
        pick = idx_map[options.index(choice)]
        tname, tid, vname, vid = views[pick]

    air_url = f"https://airtable.com/{AIRTABLE_BASE_ID}/{tid}/{vid}"
    st.markdown(f"🔗 **Lien Airtable :** [{tname} • {vname}]({air_url})")

    # --- NOUVEAU : bouton pour charger la vue sélectionnée ---
    load_votes = st.button("⚙️ Charger les données Airtable")

    if not load_votes:
        # On s'arrête ici tant que l'utilisateur n'a pas cliqué
        st.info("Sélectionne une vue puis clique sur **⚙️ Charger les données Airtable**.")
        st.stop()

    # À partir d'ici : tout le reste de ta logique inchangée
    with st.spinner("Récupération de la vue Airtable…"):
        df_view = flatten(fetch_view_records(AIRTABLE_BASE_ID, tid, vid))

    st.write(f"**Vue :** {tname} • {vname} — {len(df_view):,} lignes")
    st.dataframe(df_view, use_container_width=True)

    # 2) Résolution du projet via BO (name + department)
    st.subheader("2) Résolution du projet (référence BO name + department)")
    project_id, project_disp = pick_project_id_from_airtable(
        df_view,
        projects_df,
        vname,
        tname,
    )
    proj_url = LPB_PROJECT_URL.format(project_id=project_id)
    st.markdown(f"🔗 **Projet choisi :** {project_disp} → [{proj_url}]({proj_url})")
    # 🔎 Date résultat de vote depuis Notion (via ID Back-office)
    date_res_vote = get_vote_result_date_for_project(project_id)
    if date_res_vote:
        st.info(f"📅 **Date résultat de vote (Notion)** : {date_res_vote}")
    else:
        st.info("📅 **Date résultat de vote (Notion)** : non renseignée pour ce projet.")

    # 3) Emails & réponses Airtable (prolongation + pouvoirs)
    st.subheader("3) Emails & Réponses Airtable (prolongation / pénalités)")

    prolong_col = detect_prolongation_column(df_view)
    pouvoir_col = detect_pouvoir_column(df_view)

    if prolong_col is None and pouvoir_col is None:
        st.error(
            "❌ Cette vue ne contient ni colonne liée à la prolongation ni colonne liée aux pouvoirs.\n\n"
            "Merci de sélectionner une autre vue ou d’ajouter ces colonnes dans la vue Airtable."
        )
        st.stop()

    email_cols = detect_email_columns(df_view)
    if not email_cols:
        st.error("Impossible de détecter une colonne e-mail dans la vue (aucun '@').")
        st.stop()

    picked_em_cols = st.multiselect(
        "Colonnes e-mail à utiliser :",
        options=email_cols,
        default=email_cols[:1],
    )

    st.caption(
        f"Colonne 'prolongation' détectée : **{prolong_col or 'Aucune'}** — "
        f"colonne 'pouvoir' détectée : **{pouvoir_col or 'Aucune'}**"
    )

    if pouvoir_col is not None:
        st.markdown(
            "**Questions associées :** \n"
            "- **Prolongation** : _« Êtes-vous d'accord pour accorder la prolongation ? »_ \n"
            "- **Pouvoir (pénalités)** : _« Êtes-vous d'accord pour **NE PAS APPLIQUER** les pénalités ? »_"
        )
    else:
        st.markdown(
            "**Question associée :** \n"
            "- **Prolongation** : _« Êtes-vous d'accord pour appliquer la prolongation, "
            "avec application des 5% de pénalité ? »_"
        )

    votes_clean = build_votes_email_flags(df_view, picked_em_cols, prolong_col, pouvoir_col)

    # Doublons
    dups = votes_clean[votes_clean["n_occur"] > 1].sort_values("n_occur", ascending=False)
    st.write("Adresses en doublon (après normalisation) :")
    if dups.empty:
        st.success("Aucun doublon détecté ✅")
    else:
        st.dataframe(
            dups.rename(
                columns={"email_raw_example": "Adresse complète", "n_occur": "Nombre de doublons"}
            ),
            use_container_width=True,
        )
        st.download_button(
            "💾 Exporter les doublons (CSV)",
            data=(
                dups[["email_normalized", "email_raw_example", "n_occur"]]
                .rename(
                    columns={
                        "email_raw_example": "adresse_complete",
                        "n_occur": "nombre_doublons",
                    }
                )
                .to_csv(index=False)
                .encode("utf-8")
            ),
            file_name="adresses_doublons.csv",
            mime="text/csv",
        )

    st.write(
        f"Emails uniques dans la vue (après normalisation/déduplication) : **{len(votes_clean):,}**"
    )

    # 4) Souscriptions du projet (BO)
    st.subheader("4) Souscriptions BO du projet (filtrées par project_id)")
    subs = load_subs_for_project(project_id)

    # 🔢 Calcul du nombre d'investisseurs uniques (emails uniques)
    total_investors = (
        subs["email_normalized"]
        .dropna()
        .astype(str)
        .str.strip()
        .str.lower()
        .nunique()
    )
    
    st.write(f"Souscriptions uniques (avec e-mail) : {total_investors:,}")

    # 5) Croisement e-mail (Airtable ↔ BO)
    st.subheader("5) Croisement e-mail (Airtable ↔ BO)")
    merged = votes_clean.merge(
        subs[["email_normalized", "users_profile_id", "subscription_id", "subscribed_at"]],
        on="email_normalized",
        how="left",
    )

    # Garder la souscription la plus récente si plusieurs
    merged = (
        merged.sort_values(["email_normalized", "subscribed_at"], ascending=[True, False])
        .drop_duplicates("email_normalized")
    )

    # A. Votants sans souscription
    off_proj = merged[merged["subscription_id"].isna()].copy().sort_values("email_normalized")
    st.write("Adresses sans souscription détectée :")
    if off_proj.empty:
        st.success("Toutes les adresses ont au moins une souscription détectée ✅")
    else:
        st.dataframe(
            off_proj[["email_raw_example", "email_normalized", "prolongation", "pouvoir"]]
            .rename(columns={"email_raw_example": "Adresse complète"}),
            use_container_width=True,
        )
        st.download_button(
            "💾 Exporter les adresses sans souscription (CSV)",
            data=(
                off_proj[
                    ["email_raw_example", "email_normalized", "prolongation", "pouvoir"]
                ]
                .rename(columns={"email_raw_example": "adresse_complete"})
                .to_csv(index=False)
                .encode("utf-8")
            ),
            file_name=f"adresses_sans_souscription_project_{project_id}.csv",
            mime="text/csv",
        )

    st.markdown("---")
    st.subheader("📊 KPI résultats")

    # Table finale = emails avec souscription
    final_tbl = merged[merged["subscription_id"].notna()].copy()
    final_tbl = final_tbl[
        [
            "email_raw_example",
            "email_normalized",
            "users_profile_id",
            "subscription_id",
            "subscribed_at",
            "prolongation",
            "pouvoir",
        ]
    ].rename(columns={"email_raw_example": "Adresse complète"})

    # ===================== POIDS (€ investis) — INFORMATIONS =====================
    weights_df = load_invest_amounts_for_project(project_id)
    
    final_tbl = final_tbl.merge(weights_df, on="email_normalized", how="left")
    final_tbl["invest_amount_eur"] = final_tbl["invest_amount_eur"].fillna(0.0)


    n_votes = len(votes_clean)
    if not dups.empty:
        n_dups_total = int(dups["n_occur"].sum())
        n_dups_unique = int(len(dups))
        n_dups = int((dups["n_occur"] - 1).sum())
    else:
        n_dups_total = n_dups_unique = n_dups = 0

    n_with = final_tbl.shape[0]
    n_without = off_proj.shape[0]
    coverage_rate = (n_with / n_votes * 100) if n_votes else 0.0
    total_subs = (
        subs["email_normalized"].dropna().astype(str).str.strip().str.lower().nunique()
    )
    participation_rate = (n_with / total_subs * 100) if total_subs else 0.0
    part_delta = f"{n_with}/{total_subs}" if total_subs else "0/0 (aucune souscription projet)"
    # On réutilise / sécurise la date de résultat de vote Notion
    date_res_vote = date_res_vote if "date_res_vote" in locals() else get_vote_result_date_for_project(project_id)


    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Adresses uniques (vue)", f"{n_votes:,}")
    c2.metric("Doublons", f"{n_dups:,}")
    c3.metric("Avec souscription", f"{n_with:,}")
    c4.metric("Sans souscription", f"{n_without:,}")
    c5.metric("Taux de couverture", f"{coverage_rate:.1f}%")
    c6.metric("Participation réelle", f"{participation_rate:.1f}%", delta=part_delta)
    
    c7, _, _ = st.columns(3)
    with c7:
        st.metric("Date résultat de vote (Notion)", date_res_vote or "Non renseignée")


    st.caption(
        "KPI calculés après déduplication. "
        "‘Couverture’ = adresses uniques de la vue avec souscription. "
        "‘Participation’ = souscripteurs ayant répondu / souscripteurs totaux du projet."
    )

    # =========================== Répartition Prolongation & Pouvoirs ===========================
    st.subheader("❓ Répartition des réponses par question (votes légitimes)")

    if n_with == 0:
        st.warning(
            "Aucune souscription BO n'a pu être associée aux adresses de la vue. "
            "Il n'y a donc **aucun vote légitime** pour ce projet : "
            "les répartitions de réponses et le verdict ne sont pas calculés."
        )

        st.subheader("🧾 Verdict final")
        st.info(
            "Impossible de calculer un verdict : aucun vote légitime "
            "(0 investisseur avec souscription associée à cette vue)."
        )

        st.markdown(
            "**Table finale — après nettoyage (aucune ligne car aucune souscription associée au projet)**"
        )
        st.dataframe(final_tbl.sort_values("email_normalized"), use_container_width=True)
        st.download_button(
            "💾 Export CSV (table finale)",
            data=final_tbl.to_csv(index=False).encode("utf-8"),
            file_name=f"votes_x_subs_project_{project_id}_final.csv",
            mime="text/csv",
        )
        st.stop()

    base_df = final_tbl.copy()
    acteur_label = "les investisseurs (votes légitimes, e-mails avec souscription)"

    st.caption(
        "Les répartitions et le verdict ci-dessous sont calculés **uniquement** "
        "sur les investisseurs identifiés (votes légitimes, associés à une souscription LPB)."
    )

    def render_pie(counts, title: str):
        """Retourne une figure matplotlib pour un camembert propre."""
        fig, ax = plt.subplots()
        ax.pie(
            counts["Nombre"],
            labels=counts.index,
            autopct="%1.1f%%",
            startangle=90,
        )
        ax.set_title(title)
        ax.axis("equal")
        return fig

    # ---- Analyse Prolongation ----
    counts_pro = None
    df_pro = None
    q_pro = ""

    if prolong_col is not None and "prolongation" in base_df.columns:
        s_pro = base_df["prolongation"].fillna("Non renseigné").astype(str).str.strip()
        counts_pro = s_pro.value_counts()
        df_pro = counts_pro.reset_index()
        df_pro.columns = ["Réponse", "Nombre"]
        df_pro["%"] = (df_pro["Nombre"] / df_pro["Nombre"].sum() * 100).round(1)
        if pouvoir_col is not None:
            q_pro = "Êtes-vous d'accord pour ACCORDER la prolongation ?"
        else:
            q_pro = (
                "Êtes-vous d'accord pour APPLIQUER la prolongation, "
                "avec application des 5% de pénalités ?"
            )

    # ---- Analyse Pouvoir (Pénalités) ----
    counts_pvr = None
    df_pvr = None
    q_pvr = ""

    if pouvoir_col is not None and "pouvoir" in base_df.columns:
        s_pvr = base_df["pouvoir"].fillna("Non renseigné").astype(str).str.strip()
        counts_pvr = s_pvr.value_counts()
        df_pvr = counts_pvr.reset_index()
        df_pvr.columns = ["Réponse", "Nombre"]
        df_pvr["%"] = (df_pvr["Nombre"] / df_pvr["Nombre"].sum() * 100).round(1)
        q_pvr = "Êtes-vous d'accord pour NE PAS APPLIQUER les pénalités ?"

    col1, col2 = st.columns(2)
    if counts_pro is not None:
        with col1:
            st.markdown("### Question — Prolongation")
            st.caption(q_pro)
            st.dataframe(df_pro, use_container_width=True)
            fig_pro = render_pie(df_pro.set_index("Réponse"), "Prolongation")
            st.pyplot(fig_pro, clear_figure=True)

    if counts_pvr is not None:
        with col2:
            st.markdown("### Question — Pénalités")
            st.caption(q_pvr)
            st.dataframe(df_pvr, use_container_width=True)
            fig_pvr = render_pie(df_pvr.set_index("Réponse"), "Pénalités")
            st.pyplot(fig_pvr, clear_figure=True)

    # ===================== RÉSULTAT PRORATISÉ (INFORMATIF) =====================
    st.markdown("---")
    st.subheader("⚖️ Résultat proratisé (informatif) — pondéré par le montant investi (€)")
    
    def weighted_counts(df: pd.DataFrame, col_answer: str, col_weight: str = "invest_amount_eur"):
        s = df[col_answer].fillna("Non renseigné").astype(str).str.strip()
        tmp = pd.DataFrame({"Réponse": s, "Poids (€)": df[col_weight].fillna(0.0)})
        out = tmp.groupby("Réponse", as_index=False)["Poids (€)"].sum()
        total = float(out["Poids (€)"].sum())
        out["%"] = ((out["Poids (€)"] / total) * 100).round(1) if total > 0 else 0.0
        out = out.sort_values("Poids (€)", ascending=False)
        return out, total
    
    def render_pie_weighted(df_counts: pd.DataFrame, title: str):
        fig, ax = plt.subplots()
        ax.pie(
            df_counts["Poids (€)"],
            labels=df_counts["Réponse"],
            autopct="%1.1f%%",
            startangle=90,
        )
        ax.set_title(title)
        ax.axis("equal")
        return fig
    
    col1w, col2w = st.columns(2)
    
    # Prolongation pondérée
    if counts_pro is not None and "prolongation" in final_tbl.columns:
        with col1w:
            st.markdown("### Prolongation — pondéré")
            df_w_pro, tot_w_pro = weighted_counts(final_tbl, "prolongation")
            st.dataframe(df_w_pro, use_container_width=True)
            st.caption(f"Total pondéré : {tot_w_pro:,.2f} €")
            st.pyplot(render_pie_weighted(df_w_pro, "Prolongation (pondéré)"), clear_figure=True)
    
    # Pénalités / pouvoir pondéré
    if counts_pvr is not None and "pouvoir" in final_tbl.columns:
        with col2w:
            st.markdown("### Pénalités — pondéré")
            df_w_pvr, tot_w_pvr = weighted_counts(final_tbl, "pouvoir")
            st.dataframe(df_w_pvr, use_container_width=True)
            st.caption(f"Total pondéré : {tot_w_pvr:,.2f} €")
            st.pyplot(render_pie_weighted(df_w_pvr, "Pénalités (pondéré)"), clear_figure=True)
    

    # ================================ Verdict final ================================
    st.subheader("🧾 Verdict final")

    def get_yes_no(counts):
        """Retourne uniquement les Oui/Non sous forme (yes, no, total_exprimes)."""
        yes = int(counts.get("Oui", 0))
        no = int(counts.get("Non", 0))
        total_exprimes = yes + no
        return yes, no, total_exprimes

    verdict_parts = []

    # Verdict prolongation
    if counts_pro is not None:
        yes_p, no_p, tot_p = get_yes_no(counts_pro)
        if tot_p == 0:
            verdict_parts.append(
                "Sur la question de la prolongation, aucun vote Oui/Non exploitable n'a été exprimé."
            )
        else:
            if pouvoir_col is not None:
                # Cas avec pénalités séparées
                if yes_p > no_p:
                    verdict_parts.append(
                        f"Sur la question de la prolongation, {acteur_label} "
                        f"**ACCEPTENT la prolongation** "
                        f"({yes_p} Oui / {tot_p} votes exprimés)."
                    )
                elif no_p > yes_p:
                    verdict_parts.append(
                        f"Sur la question de la prolongation, {acteur_label} "
                        f"**REFUSENT la prolongation** "
                        f"({no_p} Non / {tot_p} votes exprimés)."
                    )
                else:
                    verdict_parts.append(
                        "Sur la question de la prolongation, il y a **égalité parfaite** "
                        f"({yes_p} Oui / {no_p} Non). Décision manuelle nécessaire."
                    )
            else:
                # Cas sans colonne pouvoir : prolongation + pénalités
                if yes_p > no_p:
                    verdict_parts.append(
                        "Sur la question « prolongation avec 5% de pénalités », "
                        f"{acteur_label} **ACCEPTENT la prolongation avec pénalités** "
                        f"({yes_p} Oui / {tot_p} votes exprimés)."
                    )
                elif no_p > yes_p:
                    verdict_parts.append(
                        "Sur la question « prolongation avec 5% de pénalités », "
                        f"{acteur_label} **REFUSENT la prolongation avec pénalités** "
                        f"({no_p} Non / {tot_p} votes exprimés)."
                    )
                else:
                    verdict_parts.append(
                        "Sur la question « prolongation avec 5% de pénalités », il y a "
                        f"**égalité parfaite** ({yes_p} Oui / {no_p} Non). "
                        "Décision manuelle nécessaire."
                    )

    # Verdict pénalités (pouvoir)
    if counts_pvr is not None:
        yes_pen, no_pen, tot_pen = get_yes_no(counts_pvr)
        if tot_pen == 0:
            verdict_parts.append(
                "Sur la question des pénalités, aucun vote Oui/Non exploitable n'a été exprimé."
            )
        else:
            if yes_pen > no_pen:
                verdict_parts.append(
                    "Sur la question des pénalités, "
                    f"{acteur_label} **VALIDENT la non-application des pénalités** "
                    f"({yes_pen} Oui / {tot_pen} votes exprimés)."
                )
            elif no_pen > yes_pen:
                verdict_parts.append(
                    "Sur la question des pénalités, "
                    f"{acteur_label} **REFUSENT la non-application des pénalités** "
                    f"({no_pen} Non / {tot_pen} votes exprimés)."
                )
            else:
                verdict_parts.append(
                    "Sur la question des pénalités, il y a **égalité parfaite** "
                    f"({yes_pen} Oui / {no_pen} Non). Décision manuelle nécessaire."
                )

    if not verdict_parts:
        st.markdown("Impossible de calculer un verdict : aucune donnée exploitable.")
    else:
        for v in verdict_parts:
            st.markdown("➡️ " + v)

    # ================================ Table finale + export ================================
    st.markdown(
        "**Table finale — après nettoyage (doublons supprimés) et retrait des invests sans souscription**"
    )
    st.dataframe(final_tbl.sort_values("email_normalized"), use_container_width=True)
    st.download_button(
        "💾 Export CSV (table finale)",
        data=final_tbl.to_csv(index=False).encode("utf-8"),
        file_name=f"votes_x_subs_project_{project_id}_final.csv",
        mime="text/csv",
    )


# =============================================================================
# 📄 MODULE 2 : PAGE "Data Hub LPB (Notion + Back-office)"
# =============================================================================

def page_data_hub():
    st.title("🛠️ Data Hub (BO/Notion)")
    st.caption(
        "Connexion Notion & Back-office. "
        "Référentiels de données : "
        "[PostgreSQL](https://lapremierebriquelpb-my.sharepoint.com/:x:/g/personal/r_taugourdeau_lapremierebrique_fr/IQBupYFLm1N4RbSbYJ-Pz8bLAfb3JD6Dj4_mQLDjyD-0Za0?e=CK5PqI) · "
        "[Notion](https://lapremierebriquelpb-my.sharepoint.com/:x:/g/personal/r_taugourdeau_lapremierebrique_fr/IQD3BNen_QP3SbgdN8wPwagIAWCviH7UbIpyaS_Eze41GIQ?e=r6Xzzt)"
    )

    # ---- Sidebar spécifique Data Hub ----
    with st.sidebar:
        section = st.radio(
            "Data Hub — Section",
            ["Notion", "Back-office"], #  ["Notion", "Back-office"] -> BO annulé finalement
            index=0,
        )

    # ============================ NOTION ============================
    if section == "Notion":
        st.subheader("🏠 Notion")
        st.caption(
            "Connexion API Notion ou import de l'export Notion (ZIP/CSV) via : "
            "https://www.notion.so/lapremierebrique/20a18ece1f2d81fab4bbf17f57df8a3a?v=21418ece1f2d8031baf2000c82651d2f"
        )

        # 👉 Drag & drop (ZIP/CSV) en premier, API en deuxième
        tab_import, tab_api = st.tabs(["Mode Export (ZIP/CSV)", "Mode API (plus long)"])

        # ---------- Mode Import ZIP/CSV ----------
        with tab_import:
            st.subheader("Import d’un export Notion (ZIP/CSV)")
            uploaded = st.file_uploader(
                "Dépose ici ton export Notion : soit le ZIP principal, soit un CSV",
                type=["zip", "csv"],
                accept_multiple_files=False,
            )

            if uploaded is not None:
                try:
                    if uploaded.name.lower().endswith(".csv"):
                        df_import = pd.read_csv(uploaded)
                        chosen_name = uploaded.name
                    else:
                        with st.spinner("Lecture du ZIP Notion (ZIP → ZIP → CSV)…"):
                            csv_list = extract_csv_recursive(uploaded)
                        if not csv_list:
                            st.error("Aucun CSV trouvé dans ce ZIP.")
                            df_import = None
                            chosen_name = None
                        else:
                            all_candidates = [
                                c for c in csv_list if c[0].lower().endswith("_all.csv")
                            ]
                            if len(all_candidates) == 1:
                                chosen_name, df_import = all_candidates[0]
                            else:
                                chosen_name = st.selectbox(
                                    "Plusieurs CSV trouvés, choisis celui à afficher :",
                                    [name for name, _ in csv_list],
                                )
                                df_import = next(
                                    df for (name, df) in csv_list if name == chosen_name
                                )

                    if df_import is not None:
                        st.session_state.imp_df = df_import
                        st.session_state.imp_filename = chosen_name
                        st.session_state.pop("imp_cols_multiselect", None)
                except Exception as e:
                    st.error(f"❌ Erreur lors de la lecture de l’export : {e}")

            df_import = st.session_state.imp_df
            chosen_name = st.session_state.imp_filename

            if df_import is None:
                st.info("Glisse ton fichier ici pour commencer.")
            else:
                st.success(
                    f"Fichier chargé : **{chosen_name}** "
                    f"({len(df_import)} lignes × {len(df_import.columns)} colonnes)"
                )

                all_cols_imp = list(df_import.columns)
                if "imp_cols_multiselect" not in st.session_state:
                    st.session_state.imp_cols_multiselect = all_cols_imp.copy()
                else:
                    st.session_state.imp_cols_multiselect = [
                        c
                        for c in st.session_state.imp_cols_multiselect
                        if c in all_cols_imp
                    ]

                st.caption("Colonnes à afficher (tu peux taper pour filtrer les noms).")
                c2_sel1, c2_sel2 = st.columns(2)
                with c2_sel1:
                    if st.button("✅ Tout sélectionner", key="imp_select_all"):
                        st.session_state.imp_cols_multiselect = all_cols_imp.copy()
                with c2_sel2:
                    if st.button("🚫 Tout désélectionner", key="imp_select_none"):
                        st.session_state.imp_cols_multiselect = []

                st.multiselect(
                    "Colonnes à afficher",
                    options=all_cols_imp,
                    key="imp_cols_multiselect",
                )

                selected_cols_imp = st.session_state.imp_cols_multiselect
                if selected_cols_imp:
                    df_view_imp = df_import[selected_cols_imp].copy()
                else:
                    df_view_imp = df_import.iloc[:, []].copy()

                edited_df_imp = st.data_editor(
                    df_view_imp,
                    use_container_width=True,
                    height=550,
                    hide_index=True,
                    key="imp_table",
                )

                # ✅ limiter export (global)
                df_export_imp = edited_df_imp.head(MAX_ROWS_EXPORT).copy()
                
                # CSV : si gros => ZIP en plusieurs CSV
                csv_is_chunked_imp = len(df_export_imp) > ROWS_PER_CSV_FILE
                if csv_is_chunked_imp:
                    zip_name_imp = f"{TODAY_STR}_notion_export_csv.zip"
                    zip_bytes_imp = chunk_df_to_zip_csv_bytes(
                        df_export_imp,
                        rows_per_file=ROWS_PER_CSV_FILE,
                        base_name=f"{TODAY_STR}_notion_export",
                    )
                else:
                    csv_name_imp = f"{TODAY_STR}_notion_export.csv"
                    csv_bytes_imp = df_to_csv_bytes(df_export_imp)
                
                # Excel : limite stricte
                df_excel_imp = df_export_imp.head(MAX_ROWS_EXCEL).copy()
                xlsx_name_imp = f"{TODAY_STR}_notion_export.xlsx"
                excel_bytes_imp = df_to_excel_bytes(df_excel_imp, sheet_name="ExportNotion")
                
                c1, c2 = st.columns(2)
                with c1:
                    if csv_is_chunked_imp:
                        st.download_button("📥 Télécharger CSV", data=zip_bytes_imp, file_name=zip_name_imp, mime="application/zip")
                    else:
                        st.download_button("📥 Télécharger CSV", data=csv_bytes_imp, file_name=csv_name_imp, mime="text/csv")
                
                with c2:
                    st.download_button("📥 Télécharger Excel", data=excel_bytes_imp, file_name=xlsx_name_imp,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
                if len(edited_df_imp) > MAX_ROWS_EXPORT:
                    st.warning(f"⚠️ Export tronqué : {MAX_ROWS_EXPORT:,} premières lignes sur {len(edited_df_imp):,}.")


        # ---------- Mode API ----------
        with tab_api:
            st.subheader("Connexion directe à Notion (API)")
            base_choisie = st.radio(
                "Sélection",
                ["Notion projet", "Bac à sable"],
                horizontal=True,
            )
            db_id = DB_BAC_SABLE if base_choisie == "Bac à sable" else DB_NOTION_PROJET

            if st.button("⚙️ Charger via l’API Notion"):
                try:
                    with st.spinner("Récupération du schéma Notion…"):
                        schema_order = get_database_schema(db_id)
                    with st.spinner("Récupération de toutes les lignes…"):
                        results = get_all_rows(db_id)
                    if not results:
                        st.info("Aucune page trouvée dans cette base.")
                        st.session_state.api_df = None
                    else:
                        df_api = notion_to_df(results, schema_order)
                        st.session_state.api_df = df_api
                        st.session_state.pop("api_cols_multiselect", None)
                except Exception as e:
                    st.error(f"❌ Erreur API Notion : {e}")

            df_api = st.session_state.api_df
            if df_api is None or df_api.empty:
                st.info(
                    "Aucune donnée API chargée pour le moment. "
                    "Clique sur le bouton ci-dessus."
                )
            else:
                st.success(f"{len(df_api)} lignes • {len(df_api.columns)} colonnes")

                all_cols = list(df_api.columns)
                if "api_cols_multiselect" not in st.session_state:
                    st.session_state.api_cols_multiselect = all_cols.copy()
                else:
                    st.session_state.api_cols_multiselect = [
                        c
                        for c in st.session_state.api_cols_multiselect
                        if c in all_cols
                    ]

                st.caption("Colonnes à afficher (tu peux taper pour filtrer les noms).")
                c_sel1, c_sel2 = st.columns(2)
                with c_sel1:
                    if st.button("✅ Tout sélectionner", key="api_select_all"):
                        st.session_state.api_cols_multiselect = all_cols.copy()
                with c_sel2:
                    if st.button("🚫 Tout désélectionner", key="api_select_none"):
                        st.session_state.api_cols_multiselect = []

                st.multiselect(
                    "Colonnes à afficher",
                    options=all_cols,
                    key="api_cols_multiselect",
                )

                selected_cols = st.session_state.api_cols_multiselect
                if selected_cols:
                    df_view = df_api[selected_cols].copy()
                else:
                    df_view = df_api.iloc[:, []].copy()

                edited_df = st.data_editor(
                    df_view,
                    use_container_width=True,
                    height=550,
                    hide_index=True,
                    key="api_table",
                )

                base_slug = base_choisie.replace(" ", "_").lower()

                # ✅ limiter export (global)
                df_export_api = edited_df.head(MAX_ROWS_EXPORT).copy()
                
                # CSV : si gros => ZIP en plusieurs CSV
                csv_is_chunked_api = len(df_export_api) > ROWS_PER_CSV_FILE
                if csv_is_chunked_api:
                    zip_name = f"{TODAY_STR}_notion_api_{base_slug}_csv.zip"
                    zip_bytes = chunk_df_to_zip_csv_bytes(
                        df_export_api,
                        rows_per_file=ROWS_PER_CSV_FILE,
                        base_name=f"{TODAY_STR}_notion_api_{base_slug}",
                    )
                else:
                    csv_name = f"{TODAY_STR}_notion_api_{base_slug}.csv"
                    csv_bytes = df_to_csv_bytes(df_export_api)
                
                # Excel : limite stricte
                df_excel_api = df_export_api.head(MAX_ROWS_EXCEL).copy()
                xlsx_name = f"{TODAY_STR}_notion_api_{base_slug}.xlsx"
                excel_bytes = df_to_excel_bytes(df_excel_api, sheet_name=base_choisie[:31])
                
                c1, c2 = st.columns(2)
                with c1:
                    if csv_is_chunked_api:
                        st.download_button("📥 Télécharger CSV", data=zip_bytes, file_name=zip_name, mime="application/zip")
                    else:
                        st.download_button("📥 Télécharger CSV", data=csv_bytes, file_name=csv_name, mime="text/csv")
                
                with c2:
                    st.download_button("📥 Télécharger Excel", data=excel_bytes, file_name=xlsx_name,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
                if len(edited_df) > MAX_ROWS_EXPORT:
                    st.warning(f"⚠️ Export tronqué : {MAX_ROWS_EXPORT:,} premières lignes sur {len(edited_df):,}.")


    # ============================ BACK-OFFICE ============================
    elif section == "Back-office":
        st.subheader("🗄 Back-office PostgreSQL")
        st.caption("Connexion en lecture au read-replica PostgreSQL de production.")

        with st.expander("ℹ️ Aide — Exports & limites", expanded=False):
            st.markdown(
                """
        **Comment fonctionnent les exports :**
        
        - Les exports sont **limités automatiquement** pour éviter les crashs (RAM Streamlit limitée).
        - **CSV** :
          - petit volume → **1 CSV**
          - gros volume → **ZIP avec plusieurs CSV** (pagination automatique)
        - **Excel** :
          - export **limité** (fichiers Excel trop gros = plantage)
        
        Si un export est tronqué, un message s’affiche.
        
        ---
        
        **Fichiers CSV paginés (ZIP)**  
        Quand un ZIP est généré, il contient plusieurs fichiers :
        `export_001.csv`, `export_002.csv`, etc.  
        ➡️ Ils peuvent être **fusionnés facilement** en local (Python / Excel PowerQuery).
        
        ---
        
        ⚠️ **Tables à éviter absolument (trop volumineuses)**  
        Ces tables peuvent faire planter Streamlit même en lecture :
        - `subscriptions`
        - `payment_operations`
        - `loans_terms`
        
        👉 Si besoin, utiliser **des filtres forts** (project_id, dates, statut)  
        ou passer par un outil SQL dédié.
                """
            )        

        # Connexion & liste des tables
        try:
            with st.spinner("Connexion à PostgreSQL et récupération des tables…"):
                tables_df = list_pg_tables()
                st.session_state.bo_tables_df = tables_df
            st.success("✅ Connexion PostgreSQL réussie")
        except Exception as e:
            st.error(f"❌ Erreur de connexion PostgreSQL : {e}")
            st.stop()

        tables_df = st.session_state.bo_tables_df
        if tables_df is None or tables_df.empty:
            st.info("Aucune table utilisateur trouvée dans cette base.")
            st.stop()

        # Optionnel : schéma image si tu as le fichier
        try:
            st.image("assets/schema_bo.png")
        except Exception:
            pass

        tables_df["full_name"] = tables_df["table_schema"] + "." + tables_df["table_name"]

        def make_label(row):
            if row["table_schema"] == "public":
                return row["table_name"]
            else:
                return row["table_schema"] + "." + row["table_name"]

        tables_df["label"] = tables_df.apply(make_label, axis=1)
        table_labels = tables_df["label"].tolist()
        full_names = tables_df["full_name"].tolist()
        label_to_fullname = dict(zip(table_labels, full_names))

        default_full = st.session_state.bo_table_id or full_names[0]
        if default_full in full_names:
            default_label = tables_df.loc[
                tables_df["full_name"] == default_full, "label"
            ].iloc[0]
        else:
            default_label = table_labels[0]

        selected_label = st.selectbox(
            "Table PostgreSQL",
            options=table_labels,
            index=table_labels.index(default_label),
            help="Tu peux taper ici pour filtrer les tables en temps réel.",
        )

        selected_full = label_to_fullname[selected_label]
        st.session_state.bo_table_id = selected_full
        schema, table = selected_full.split(".", 1)

        # --- boutons côte à côte (Charger / Rafraîchir) ---
        b_load, b_refresh = st.columns([2, 1])
        
        with b_load:
            load_clicked = st.button("⚙️ Charger via Postgre", use_container_width=True)
        
        with b_refresh:
            if st.button("🔄 Rafraîchir", use_container_width=True):
                st.rerun()  # équivalent Ctrl+R
        
        if load_clicked:
            try:
                with st.spinner(f"Lecture de {schema}.{table}"):
                    df_bo = read_pg_table(schema, table)
                st.session_state.bo_df = df_bo
                st.session_state.pop("bo_cols_multiselect", None)
            except Exception as e:
                st.error(f"❌ Erreur lors de la lecture de la table : {e}")


        df_bo = st.session_state.bo_df
        if df_bo is None or df_bo.empty:
            st.info(
                "Aucune donnée chargée pour le moment. "
                "Choisis une table puis clique sur le bouton."
            )
        else:
            st.success(f"{len(df_bo)} lignes chargées • {len(df_bo.columns)} colonnes")

            all_cols_bo = list(df_bo.columns)
            if "bo_cols_multiselect" not in st.session_state:
                st.session_state.bo_cols_multiselect = all_cols_bo.copy()
            else:
                st.session_state.bo_cols_multiselect = [
                    c
                    for c in st.session_state.bo_cols_multiselect
                    if c in all_cols_bo
                ]

            st.caption("Colonnes à afficher (tu peux taper pour filtrer les noms).")
            b1, b2 = st.columns(2)
            with b1:
                if st.button("✅ Tout sélectionner", key="bo_select_all"):
                    st.session_state.bo_cols_multiselect = all_cols_bo.copy()
            with b2:
                if st.button("🚫 Tout désélectionner", key="bo_select_none"):
                    st.session_state.bo_cols_multiselect = []

            st.multiselect(
                "Colonnes à afficher",
                options=all_cols_bo,
                key="bo_cols_multiselect",
            )

            selected_cols_bo = st.session_state.bo_cols_multiselect
            if selected_cols_bo:
                df_bo_view = df_bo[selected_cols_bo].copy()
            else:
                df_bo_view = df_bo.iloc[:, []].copy()

            edited_bo = st.data_editor(
                df_bo_view,
                use_container_width=True,
                height=550,
                hide_index=True,
                key="bo_table",
            )

            safe_table_name = selected_label.replace(".", "_")

            # ✅ limiter export (global)
            df_export_bo = edited_bo.head(MAX_ROWS_EXPORT).copy()
            
            # CSV : si gros => ZIP en plusieurs CSV
            csv_is_chunked_bo = len(df_export_bo) > ROWS_PER_CSV_FILE
            if csv_is_chunked_bo:
                zip_name_bo = f"{TODAY_STR}_{safe_table_name}_csv.zip"
                zip_bytes_bo = chunk_df_to_zip_csv_bytes(
                    df_export_bo,
                    rows_per_file=ROWS_PER_CSV_FILE,
                    base_name=f"{TODAY_STR}_{safe_table_name}",
                )
            else:
                csv_name_bo = f"{TODAY_STR}_{safe_table_name}.csv"
                csv_bytes_bo = df_to_csv_bytes(df_export_bo)
            
            # Excel : limite stricte
            df_excel_bo = df_export_bo.head(MAX_ROWS_EXCEL).copy()
            xlsx_name_bo = f"{TODAY_STR}_{safe_table_name}.xlsx"
            excel_bytes_bo = df_to_excel_bytes(df_excel_bo, sheet_name=safe_table_name[:31])
            
            c1, c2 = st.columns(2)
            with c1:
                if csv_is_chunked_bo:
                    st.download_button("📥 Télécharger CSV", data=zip_bytes_bo, file_name=zip_name_bo, mime="application/zip")
                else:
                    st.download_button("📥 Télécharger CSV", data=csv_bytes_bo, file_name=csv_name_bo, mime="text/csv")
            
            with c2:
                st.download_button("📥 Télécharger Excel", data=excel_bytes_bo, file_name=xlsx_name_bo,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            if len(edited_bo) > MAX_ROWS_EXPORT:
                st.warning(f"⚠️ Export tronqué : {MAX_ROWS_EXPORT:,} premières lignes sur {len(edited_bo):,}.")


# =============================================================================
# 📄 MODULE 3 : PAGE "Suivi Invest TDF"
# =============================================================================

OUTPUT_COLUMNS_TDF = [
    "EVENT", "DATE", "EMAIL", "NOM", "PRENOM", "INVEST",
    "CREATION COMPTE", "ENCOURS",
    "DERNIER INVEST AVANT SOIREE", "MONTANT",
    "INVEST APRES SOIREE", "MONTANT",
    "DEVENU INVEST",
    "VARIATION INVEST AV VS AP SOIREE",
    "MONTANT INVEST AV", "MONTANT INVEST AP",
]

ROW_KEYS_TDF = [
    "EVENT", "DATE", "EMAIL", "NOM", "PRENOM", "INVEST",
    "CREATION COMPTE", "ENCOURS",
    "DERNIER INVEST AVANT SOIREE", "MONTANT",
    "INVEST APRES SOIREE", "MONTANT_2",
    "DEVENU INVEST",
    "VARIATION INVEST AV VS AP SOIREE",
    "MONTANT INVEST AV", "MONTANT INVEST AP",
]

USER_NOT_FOUND_TDF = "USER NON TROUVÉ"

COLOR_HEADER_BG_TDF  = "1F3864"
COLOR_HEADER_FG_TDF  = "FFFFFF"
COLOR_OK_BG_TDF      = "C6EFCE"
COLOR_OK_FG_TDF      = "006100"
COLOR_KO_BG_TDF      = "FFC7CE"
COLOR_KO_FG_TDF      = "9C0006"
COLOR_NEUTRAL_BG_TDF = "FFEB9C"
COLOR_NEUTRAL_FG_TDF = "9C5700"
COLOR_OK_STRONG_BG_TDF = "63BE7B"
COLOR_KO_STRONG_BG_TDF = "F8696B"

ENRICH_SQL_TDF = """
WITH attendees AS (
    SELECT
        UNNEST(%(emails)s::text[]) AS email_norm,
        UNNEST(%(dates)s::date[])  AS event_date
),
matched_users AS (
    SELECT
        a.email_norm, a.event_date,
        u.id          AS user_id,
        u.first_name, u.last_name,
        u.created_at  AS creation_compte
    FROM attendees a
    LEFT JOIN users u ON LOWER(TRIM(u.email)) = a.email_norm
),
user_subs AS (
    SELECT
        m.email_norm, m.event_date, m.user_id,
        m.first_name, m.last_name, m.creation_compte,
        s.created_at AS sub_created_at,
        s.amount,
        s.status
    FROM matched_users m
    LEFT JOIN users_profiles up ON up.user_id = m.user_id
    LEFT JOIN subscriptions  s  ON s.users_profile_id = up.id
                                AND s.status <> 'canceled'
)
SELECT
    email_norm,
    event_date,
    user_id,
    first_name,
    last_name,
    creation_compte,

    CASE WHEN COUNT(*) FILTER (WHERE sub_created_at::date < event_date) > 0
         THEN 'OUI' ELSE 'NON' END                                       AS invest,

    COALESCE(SUM(amount) FILTER (WHERE status <> 'refunded'),
             0)::numeric                                                 AS encours,

    MAX(sub_created_at) FILTER (WHERE sub_created_at::date < event_date) AS dernier_invest_avant,
    (array_agg(amount ORDER BY sub_created_at DESC)
        FILTER (WHERE sub_created_at::date < event_date))[1]             AS montant_avant,

    MIN(sub_created_at) FILTER (WHERE sub_created_at::date > event_date) AS invest_apres,
    (array_agg(amount ORDER BY sub_created_at ASC)
        FILTER (WHERE sub_created_at::date > event_date))[1]             AS montant_apres,

    COALESCE(SUM(amount)
        FILTER (WHERE sub_created_at::date < event_date),
        0)::numeric                                                      AS montant_total_avant,
    COALESCE(SUM(amount)
        FILTER (WHERE sub_created_at::date > event_date),
        0)::numeric                                                      AS montant_total_apres
FROM user_subs
GROUP BY email_norm, event_date, user_id,
         first_name, last_name, creation_compte;
"""

def open_connection_tdf():
    return psycopg2.connect(
        host=PGHOST,
        port=PGPORT,
        dbname=PGDATABASE,
        user=PGUSER,
        password=PGPASSWORD,
        sslmode=PGSSLMODE,
        connect_timeout=10,
    )

def normalize_email_tdf(e):
    if e is None or (isinstance(e, float) and pd.isna(e)):
        return None
    return str(e).strip().lower()

def to_py_date_tdf(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None

def fmt_date_or_x_tdf(x):
    return x if x is not None else "X"

def fmt_amount_or_x_tdf(x):
    if x is None:
        return "X"
    try:
        return float(x)
    except (TypeError, ValueError):
        return "X"

def compute_variation_tdf(av, ap):
    try:
        av_f, ap_f = float(av), float(ap)
        if av_f == 0:
            return None
        return (ap_f - av_f) / av_f
    except (TypeError, ValueError):
        return None

def devenu_invest_flag_tdf(invest_avant, invest_apres_date):
    if invest_avant == "OUI":
        return None
    return "OUI" if invest_apres_date is not None else "NON"

def fetch_enrichment_tdf(conn, attendees_df: pd.DataFrame) -> pd.DataFrame:
    if attendees_df.empty:
        return pd.DataFrame()

    emails = attendees_df["email_norm"].tolist()
    dates = attendees_df["event_date"].tolist()

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(ENRICH_SQL_TDF, {"emails": emails, "dates": dates})
        return pd.DataFrame(cur.fetchall())

def build_output_row_tdf(input_row: pd.Series, enrich_row, event_name) -> dict:
    if enrich_row is None or pd.isna(enrich_row.get("user_id")):
        return {
            "EVENT": event_name or input_row.get("EVENT"),
            "DATE": input_row.get("DATE"),
            "EMAIL": input_row.get("EMAIL"),
            "NOM": None, "PRENOM": None,
            "INVEST": USER_NOT_FOUND_TDF,
            "CREATION COMPTE": None, "ENCOURS": None,
            "DERNIER INVEST AVANT SOIREE": None, "MONTANT": None,
            "INVEST APRES SOIREE": None, "MONTANT_2": None,
            "DEVENU INVEST": None,
            "VARIATION INVEST AV VS AP SOIREE": None,
            "MONTANT INVEST AV": None, "MONTANT INVEST AP": None,
        }

    invest = enrich_row.get("invest")
    dernier_avant = enrich_row.get("dernier_invest_avant")
    invest_apres = enrich_row.get("invest_apres")
    montant_avant = enrich_row.get("montant_avant")
    montant_apres = enrich_row.get("montant_apres")
    total_avant = enrich_row.get("montant_total_avant")
    total_apres = enrich_row.get("montant_total_apres")
    encours = enrich_row.get("encours")

    nom = enrich_row.get("last_name")
    prenom = enrich_row.get("first_name")
    nom = nom.strip().upper() if isinstance(nom, str) else nom
    prenom = prenom.strip().title() if isinstance(prenom, str) else prenom

    return {
        "EVENT": event_name or input_row.get("EVENT"),
        "DATE": input_row.get("DATE"),
        "EMAIL": input_row.get("EMAIL"),
        "NOM": nom,
        "PRENOM": prenom,
        "INVEST": invest,
        "CREATION COMPTE": enrich_row.get("creation_compte"),
        "ENCOURS": float(encours) if encours is not None else None,
        "DERNIER INVEST AVANT SOIREE": fmt_date_or_x_tdf(dernier_avant),
        "MONTANT": fmt_amount_or_x_tdf(montant_avant),
        "INVEST APRES SOIREE": fmt_date_or_x_tdf(invest_apres),
        "MONTANT_2": fmt_amount_or_x_tdf(montant_apres),
        "DEVENU INVEST": devenu_invest_flag_tdf(invest, invest_apres),
        "VARIATION INVEST AV VS AP SOIREE": compute_variation_tdf(montant_avant, montant_apres),
        "MONTANT INVEST AV": float(total_avant) if total_avant is not None else None,
        "MONTANT INVEST AP": float(total_apres) if total_apres is not None else None,
    }

THIN_TDF = Side(style="thin", color="BFBFBF")
BORDER_TDF = Border(left=THIN_TDF, right=THIN_TDF, top=THIN_TDF, bottom=THIN_TDF)
PCT_COLS_TDF = {"VARIATION INVEST AV VS AP SOIREE"}
MONEY_COLS_TDF = {"ENCOURS", "MONTANT INVEST AV", "MONTANT INVEST AP"}

def _style_header_cell_tdf(cell):
    cell.font = Font(bold=True, color=COLOR_HEADER_FG_TDF, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG_TDF)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_TDF

def _color_variation_tdf(cell, value):
    if value is None or not isinstance(value, (int, float)):
        return
    if value > 0.5:
        cell.fill = PatternFill("solid", fgColor=COLOR_OK_STRONG_BG_TDF)
        cell.font = Font(bold=True, color="FFFFFF")
    elif value > 0:
        cell.fill = PatternFill("solid", fgColor=COLOR_OK_BG_TDF)
        cell.font = Font(color=COLOR_OK_FG_TDF, bold=True)
    elif value < -0.5:
        cell.fill = PatternFill("solid", fgColor=COLOR_KO_STRONG_BG_TDF)
        cell.font = Font(bold=True, color="FFFFFF")
    elif value < 0:
        cell.fill = PatternFill("solid", fgColor=COLOR_KO_BG_TDF)
        cell.font = Font(color=COLOR_KO_FG_TDF, bold=True)

def _apply_cell_format_tdf(cell, key, col_name, value):
    cell.border = BORDER_TDF
    cell.alignment = Alignment(vertical="center", horizontal="left")

    if isinstance(value, (datetime, date)):
        cell.number_format = "DD/MM/YYYY"
        cell.alignment = Alignment(vertical="center", horizontal="center")
    elif col_name in PCT_COLS_TDF and isinstance(value, (int, float)):
        cell.number_format = "+0.00%;-0.00%;0.00%"
        cell.alignment = Alignment(vertical="center", horizontal="center")
        _color_variation_tdf(cell, value)
    elif (col_name in MONEY_COLS_TDF or key in {"MONTANT", "MONTANT_2"}) and isinstance(value, (int, float)):
        cell.number_format = "#,##0 €"
        cell.alignment = Alignment(vertical="center", horizontal="right")

    if col_name == "INVEST":
        cell.alignment = Alignment(vertical="center", horizontal="center")
        if value == "OUI":
            cell.fill = PatternFill("solid", fgColor=COLOR_OK_BG_TDF)
            cell.font = Font(bold=True, color=COLOR_OK_FG_TDF)
        elif value == "NON":
            cell.fill = PatternFill("solid", fgColor=COLOR_NEUTRAL_BG_TDF)
            cell.font = Font(bold=True, color=COLOR_NEUTRAL_FG_TDF)
        elif value == USER_NOT_FOUND_TDF:
            cell.fill = PatternFill("solid", fgColor=COLOR_KO_BG_TDF)
            cell.font = Font(bold=True, color=COLOR_KO_FG_TDF)

    elif col_name == "DEVENU INVEST":
        cell.alignment = Alignment(vertical="center", horizontal="center")
        if value == "OUI":
            cell.fill = PatternFill("solid", fgColor=COLOR_OK_STRONG_BG_TDF)
            cell.font = Font(bold=True, color="FFFFFF")
        elif value == "NON":
            cell.fill = PatternFill("solid", fgColor=COLOR_NEUTRAL_BG_TDF)
            cell.font = Font(color=COLOR_NEUTRAL_FG_TDF)

def _set_column_widths_tdf(ws):
    widths = {
        "EVENT": 14, "DATE": 12, "EMAIL": 32, "NOM": 18, "PRENOM": 14,
        "INVEST": 14, "CREATION COMPTE": 14, "ENCOURS": 13,
        "DERNIER INVEST AVANT SOIREE": 16, "MONTANT": 11,
        "INVEST APRES SOIREE": 16,
        "DEVENU INVEST": 13,
        "VARIATION INVEST AV VS AP SOIREE": 14,
        "MONTANT INVEST AV": 14, "MONTANT INVEST AP": 14,
    }
    for c_idx, col_name in enumerate(OUTPUT_COLUMNS_TDF, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col_name, 14)

def _write_event_sheet_tdf(wb, sheet_name: str, rows: list[dict]):
    ws = wb.create_sheet(title=sheet_name[:31])

    headers_for_sheet = [
        "EVENT", "DATE", "EMAIL", "NOM", "PRENOM", "INVEST",
        "CREATION COMPTE", "ENCOURS",
        "DERNIER INVEST AVANT SOIREE", "MONTANT (av.)",
        "INVEST APRES SOIREE", "MONTANT (apr.)",
        "DEVENU INVEST",
        "VARIATION INVEST AV VS AP SOIREE",
        "MONTANT INVEST AV", "MONTANT INVEST AP",
    ]

    for c_idx, col_name in enumerate(headers_for_sheet, start=1):
        _style_header_cell_tdf(ws.cell(row=1, column=c_idx, value=col_name))

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _) in enumerate(zip(ROW_KEYS_TDF, OUTPUT_COLUMNS_TDF), start=1):
            v = row.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            display_name = headers_for_sheet[c_idx - 1]
            _apply_cell_format_tdf(cell, key, display_name, v)

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "D2"
    _set_column_widths_tdf(ws)

    if rows:
        last_col = get_column_letter(len(headers_for_sheet))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

def _write_recap_sheet_tdf(wb, sheets_data: dict):
    ws = wb.create_sheet(title="📊 Récap", index=0)

    headers = [
        "Événement", "Date", "Présents", "Emails matchés",
        "Déjà investisseurs", "Devenus investisseurs",
        "Taux conversion", "Montant total AV", "Montant total AP",
        "Δ Montant",
    ]
    for c_idx, h in enumerate(headers, start=1):
        _style_header_cell_tdf(ws.cell(row=1, column=c_idx, value=h))

    r = 2
    for sname, rows in sheets_data.items():
        n_total = len(rows)
        n_matched = sum(1 for x in rows if x["INVEST"] != USER_NOT_FOUND_TDF)
        n_invest_av = sum(1 for x in rows if x["INVEST"] == "OUI")
        n_devenu = sum(1 for x in rows if x["DEVENU INVEST"] == "OUI")
        non_invest_av = sum(1 for x in rows if x["INVEST"] == "NON")
        conv_rate = (n_devenu / non_invest_av) if non_invest_av else None
        total_av = sum((x["MONTANT INVEST AV"] or 0) for x in rows)
        total_ap = sum((x["MONTANT INVEST AP"] or 0) for x in rows)
        delta = total_ap - total_av
        date_event = next((x["DATE"] for x in rows if x.get("DATE")), None)

        values = [sname, date_event, n_total, n_matched, n_invest_av, n_devenu, conv_rate, total_av, total_ap, delta]

        for c_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c_idx, value=v)
            cell.border = BORDER_TDF
            cell.alignment = Alignment(vertical="center", horizontal="center")

            if isinstance(v, (date, datetime)):
                cell.number_format = "DD/MM/YYYY"
            elif c_idx == 7 and isinstance(v, (int, float)):
                cell.number_format = "0.0%"
                _color_variation_tdf(cell, v)
            elif c_idx in (8, 9, 10) and isinstance(v, (int, float)):
                cell.number_format = "#,##0 €"
                cell.alignment = Alignment(vertical="center", horizontal="right")
        r += 1

    ws.row_dimensions[1].height = 38
    widths = [22, 12, 11, 14, 18, 19, 14, 17, 17, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def write_workbook_tdf(sheets_data: dict[str, list[dict]]) -> BytesIO:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sname, rows in sheets_data.items():
        _write_event_sheet_tdf(wb, sname, rows)

    _write_recap_sheet_tdf(wb, sheets_data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _style_preview_df_tdf(df: pd.DataFrame):
    def color_invest(val):
        if val == "OUI":
            return f"background-color: #{COLOR_OK_BG_TDF}; color: #{COLOR_OK_FG_TDF}; font-weight: bold"
        if val == "NON":
            return f"background-color: #{COLOR_NEUTRAL_BG_TDF}; color: #{COLOR_NEUTRAL_FG_TDF}; font-weight: bold"
        if val == USER_NOT_FOUND_TDF:
            return f"background-color: #{COLOR_KO_BG_TDF}; color: #{COLOR_KO_FG_TDF}; font-weight: bold"
        return ""

    def color_devenu(val):
        if val == "OUI":
            return f"background-color: #{COLOR_OK_STRONG_BG_TDF}; color: white; font-weight: bold"
        if val == "NON":
            return f"background-color: #{COLOR_NEUTRAL_BG_TDF}; color: #{COLOR_NEUTRAL_FG_TDF}"
        return ""

    def color_variation(val):
        if val is None or pd.isna(val):
            return ""
        if val > 0.5:
            return f"background-color: #{COLOR_OK_STRONG_BG_TDF}; color: white; font-weight: bold"
        if val > 0:
            return f"background-color: #{COLOR_OK_BG_TDF}; color: #{COLOR_OK_FG_TDF}; font-weight: bold"
        if val < -0.5:
            return f"background-color: #{COLOR_KO_STRONG_BG_TDF}; color: white; font-weight: bold"
        if val < 0:
            return f"background-color: #{COLOR_KO_BG_TDF}; color: #{COLOR_KO_FG_TDF}; font-weight: bold"
        return ""

    styler = df.style
    if "INVEST" in df.columns:
        styler = styler.map(color_invest, subset=["INVEST"])
    if "DEVENU INVEST" in df.columns:
        styler = styler.map(color_devenu, subset=["DEVENU INVEST"])
    var_col = "VARIATION INVEST AV VS AP SOIREE"
    if var_col in df.columns:
        styler = styler.map(color_variation, subset=[var_col])
        styler = styler.format({var_col: lambda x: "" if pd.isna(x) else f"{x:+.1%}"})

    for c in ("ENCOURS", "MONTANT INVEST AV", "MONTANT INVEST AP", "MONTANT (av.)", "MONTANT (apr.)"):
        if c in df.columns:
            styler = styler.format({
                c: lambda x: "" if pd.isna(x) or x == "X"
                else (f"{x:,.0f} €".replace(",", " ") if isinstance(x, (int, float)) else x)
            })
    return styler

def style_recap_tdf(df: pd.DataFrame):
    def color_conv(v):
        if v is None or pd.isna(v):
            return ""
        if v >= 0.3:
            return f"background-color: #{COLOR_OK_STRONG_BG_TDF}; color: white; font-weight: bold"
        if v > 0:
            return f"background-color: #{COLOR_OK_BG_TDF}; color: #{COLOR_OK_FG_TDF}; font-weight: bold"
        return f"background-color: #{COLOR_NEUTRAL_BG_TDF}; color: #{COLOR_NEUTRAL_FG_TDF}"

    def color_delta(v):
        if v is None or pd.isna(v):
            return ""
        if v > 0:
            return f"background-color: #{COLOR_OK_BG_TDF}; color: #{COLOR_OK_FG_TDF}; font-weight: bold"
        if v < 0:
            return f"background-color: #{COLOR_KO_BG_TDF}; color: #{COLOR_KO_FG_TDF}; font-weight: bold"
        return ""

    return (
        df.style
        .map(color_conv, subset=["Conversion"])
        .map(color_delta, subset=["Δ (€)"])
        .format({
            "Conversion": lambda x: "" if pd.isna(x) else f"{x:.1%}",
            "Total AV (€)": lambda x: f"{x:,.0f} €".replace(",", " "),
            "Total AP (€)": lambda x: f"{x:,.0f} €".replace(",", " "),
            "Δ (€)": lambda x: f"{x:+,.0f} €".replace(",", " "),
            "Date": lambda x: x.strftime("%d/%m/%Y") if isinstance(x, (date, datetime)) else "",
        })
    )

def page_suivi_invest_tdf():
    st.title("📊 Suivi Invest TDF")
    st.caption(
        "Charge un xlsx où **chaque feuille = un événement** "
        "avec les colonnes `EVENT`, `DATE`, `EMAIL`. Le reste est rempli "
        "automatiquement depuis PostgreSQL."
    )

    uploaded = st.file_uploader("📥 Fichier d'émargement (.xlsx)", type=["xlsx"], key="tdf_uploader")

    if uploaded is None:
        st.info("⬆️ Charge un fichier xlsx pour commencer.")
        return

    try:
        sheets = pd.read_excel(uploaded, sheet_name=None, dtype={"EMAIL": str})
    except Exception as e:
        st.error(f"Lecture du fichier impossible : {e}")
        return

    st.success(f"✅ {len(sheets)} feuille(s) détectée(s) : **{', '.join(sheets.keys())}**")

    with st.expander("👀 Aperçu des feuilles d'entrée", expanded=False):
        for sname, sdf in sheets.items():
            st.write(f"**{sname}** — {len(sdf)} ligne(s)")
            st.dataframe(sdf.head(8), use_container_width=True, hide_index=True)

    if not st.button("🚀 Lancer l'enrichissement", type="primary", use_container_width=True, key="tdf_launch"):
        return

    try:
        conn = open_connection_tdf()
    except Exception as e:
        st.error(f"❌ Connexion DB échouée : {e}")
        return

    sheets_data = {}
    not_found_global = []
    n = len(sheets)
    progress = st.progress(0.0, text="Préparation...")

    try:
        for i, (sname, sdf) in enumerate(sheets.items()):
            progress.progress(i / n if n else 1.0, text=f"Traitement : {sname}")

            sdf = sdf.rename(columns={c: str(c).strip().upper() for c in sdf.columns})

            required = {"EVENT", "DATE", "EMAIL"}
            missing = required - set(sdf.columns)
            if missing:
                st.warning(f"[{sname}] Colonnes manquantes : {missing} — feuille ignorée")
                continue

            sdf = sdf.dropna(subset=["EMAIL"]).copy()
            sdf["email_norm"] = sdf["EMAIL"].apply(normalize_email_tdf)
            sdf["event_date"] = sdf["DATE"].apply(to_py_date_tdf)
            sdf = sdf.dropna(subset=["email_norm", "event_date"])

            if sdf.empty:
                continue

            attendees = sdf[["email_norm", "event_date"]].drop_duplicates()

            try:
                enriched = fetch_enrichment_tdf(conn, attendees)
            except Exception as e:
                st.error(f"[{sname}] Erreur SQL : {e}")
                continue

            enriched_dict = {}
            if not enriched.empty:
                for _, r in enriched.iterrows():
                    enriched_dict[(r["email_norm"], r["event_date"])] = r

            event_name = sdf["EVENT"].iloc[0] if len(sdf) else None

            rows = []
            for _, irow in sdf.iterrows():
                er = enriched_dict.get((irow["email_norm"], irow["event_date"]))
                row = build_output_row_tdf(irow, er, event_name)
                rows.append(row)
                if row["INVEST"] == USER_NOT_FOUND_TDF:
                    not_found_global.append({
                        "Feuille": sname,
                        "EVENT": row["EVENT"],
                        "DATE": row["DATE"],
                        "EMAIL": row["EMAIL"],
                    })

            sheets_data[sname] = rows
    finally:
        conn.close()

    progress.progress(1.0, text="Génération du fichier...")
    out_buf = write_workbook_tdf(sheets_data)
    progress.empty()

    if not sheets_data:
        st.error("Aucune donnée enrichie produite.")
        return

    st.divider()
    st.header("📈 Dashboard")

    total_rows = sum(len(v) for v in sheets_data.values())
    total_match = sum(1 for v in sheets_data.values() for r in v if r["INVEST"] != USER_NOT_FOUND_TDF)
    total_invest = sum(1 for v in sheets_data.values() for r in v if r["INVEST"] == "OUI")
    total_devenu = sum(1 for v in sheets_data.values() for r in v if r["DEVENU INVEST"] == "OUI")
    total_av = sum((r["MONTANT INVEST AV"] or 0) for v in sheets_data.values() for r in v)
    total_ap = sum((r["MONTANT INVEST AP"] or 0) for v in sheets_data.values() for r in v)
    non_invest_av_total = sum(1 for v in sheets_data.values() for r in v if r["INVEST"] == "NON")
    conv_global = (total_devenu / non_invest_av_total) if non_invest_av_total else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("👥 Présents", f"{total_rows}")
    c2.metric("✉️ Emails matchés", f"{total_match}",
              delta=f"-{total_rows - total_match} non trouvés" if total_rows != total_match else None,
              delta_color="inverse")
    c3.metric("💼 Déjà investisseurs", f"{total_invest}")
    c4.metric("🎯 Devenus investisseurs", f"{total_devenu}",
              delta=f"{conv_global:.1%} de conversion" if conv_global else None,
              delta_color="normal")
    c5.metric("💰 Δ Montant", f"{total_ap - total_av:,.0f} €".replace(",", " "),
              delta=f"AP {total_ap:,.0f} € vs AV {total_av:,.0f} €".replace(",", " "),
              delta_color="off")

    st.subheader("📋 Récap par événement")

    recap_rows = []
    for sname, rows in sheets_data.items():
        n_total = len(rows)
        n_matched = sum(1 for x in rows if x["INVEST"] != USER_NOT_FOUND_TDF)
        n_invest_av = sum(1 for x in rows if x["INVEST"] == "OUI")
        n_devenu = sum(1 for x in rows if x["DEVENU INVEST"] == "OUI")
        non_invest_av = sum(1 for x in rows if x["INVEST"] == "NON")
        conv_rate = (n_devenu / non_invest_av) if non_invest_av else None
        av = sum((x["MONTANT INVEST AV"] or 0) for x in rows)
        ap = sum((x["MONTANT INVEST AP"] or 0) for x in rows)
        date_event = next((x["DATE"] for x in rows if x.get("DATE")), None)
        recap_rows.append({
            "Événement": sname,
            "Date": date_event,
            "Présents": n_total,
            "Matchés": n_matched,
            "Déjà invest.": n_invest_av,
            "Devenus invest.": n_devenu,
            "Conversion": conv_rate,
            "Total AV (€)": av,
            "Total AP (€)": ap,
            "Δ (€)": ap - av,
        })

    df_recap = pd.DataFrame(recap_rows)
    st.dataframe(style_recap_tdf(df_recap), use_container_width=True, hide_index=True)

    st.subheader("📊 Visualisations")

    col_g1, col_g2 = st.columns(2)

    with col_g1:
        st.markdown("**Conversion par événement**")
        chart_df = df_recap.set_index("Événement")[["Déjà invest.", "Devenus invest."]]
        st.bar_chart(chart_df, height=300)

    with col_g2:
        st.markdown("**Montants investis par événement (€)**")
        money_df = df_recap.set_index("Événement")[["Total AV (€)", "Total AP (€)"]]
        st.bar_chart(money_df, height=300)

    st.divider()
    st.header("🔍 Détail par événement")

    tabs = st.tabs(list(sheets_data.keys()))
    for tab, (sname, rows) in zip(tabs, sheets_data.items()):
        with tab:
            df_view = pd.DataFrame(rows)
            if "MONTANT_2" in df_view.columns:
                df_view = df_view.rename(columns={
                    "MONTANT": "MONTANT (av.)",
                    "MONTANT_2": "MONTANT (apr.)",
                })

            n_total = len(rows)
            n_matched = sum(1 for x in rows if x["INVEST"] != USER_NOT_FOUND_TDF)
            n_invest_av = sum(1 for x in rows if x["INVEST"] == "OUI")
            n_devenu = sum(1 for x in rows if x["DEVENU INVEST"] == "OUI")
            non_invest_av = sum(1 for x in rows if x["INVEST"] == "NON")
            conv = (n_devenu / non_invest_av) if non_invest_av else 0

            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric("Présents", n_total)
            mc2.metric("Matchés", n_matched)
            mc3.metric("Déjà invest.", n_invest_av)
            mc4.metric("Devenus invest.", n_devenu, delta=f"{conv:.1%}" if conv else None)

            st.dataframe(_style_preview_df_tdf(df_view), use_container_width=True, hide_index=True)

    st.divider()
    st.header("💾 Téléchargements")

    dl1, dl2 = st.columns([2, 1])

    with dl1:
        st.download_button(
            "📥 Fichier complété (xlsx)",
            data=out_buf,
            file_name=f"Suivi_Invest_TDF_complete_{datetime.now():%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )

    with dl2:
        if not_found_global:
            df_nf = pd.DataFrame(not_found_global)
            st.download_button(
                f"⚠️ {len(not_found_global)} non trouvés (CSV)",
                data=df_nf.to_csv(index=False).encode("utf-8"),
                file_name=f"non_trouves_{datetime.now():%Y%m%d_%H%M}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    if not_found_global:
        with st.expander(f"⚠️ Liste des {len(not_found_global)} email(s) non trouvé(s) en DB", expanded=False):
            st.dataframe(pd.DataFrame(not_found_global), use_container_width=True, hide_index=True)
# =============================================================================
# 📄 MODULE 3/4 : Autres pages
# =============================================================================

def page_file_splitter():
    st.title("✂️ Découpe et recollement de fichiers")
    st.markdown(
        """
        Outil de découpe et recollement de fichiers.

        👉 **Ouvrir l'application Streamlit :**  
        [https://file-splitter.streamlit.app/](https://file-splitter.streamlit.app/)
        """
    )
    
def page_prequalification():
    st.title("📁 Préqualification des projets, emailing de refus et scoring ML de la prédiction du statut")
    st.markdown(
        """
        Cette app publique est dédiée à l'équipe projet.

        👉 **Ouvrir l'application Streamlit :**  
        [https://lpb-pdp-preselection-emaling.streamlit.app/](https://lpb-pdp-preselection-emaling.streamlit.app/)
        """
    )

def page_rendements():
    st.title("📈 Comparaison des rendements par classe d'actifs")

    st.markdown(
        """
        Dashboard interactif comparant le rendement du crowdfunding immobilier (≈ 11,5 % annualisé)
        aux autres classes d'actifs, avec une étude des corrélations et une illustration de l’effet
        des **intérêts composés** (intérêts cumulés), à l’aide d’un simulateur basé sur l’ancienne
        flat tax annuelle à 30 %, à titre illustratif.

        👉 **Ouvrir l'application RShiny :**  
        https://rtaugourdeau-lpb.shinyapps.io/RendementActifs/

        👉 **Simulateur web d'intérêts composés :**  
        https://businesstracker.netlify.app/simulateurinteretscomposes  
        """
    )

def page_docs():
    st.title("📄 Documentation")

    DOCS = {
        "📊 Listing des dashboards BI":
            "https://docs.google.com/document/d/129lJRvAsvrxj4P2L_bN5zfhMeF5aXm0YV5wSJ6yLOT8/edit?tab=t.0",

        "🗄️ Architecture Fabric & MODOP Power BI":
            "https://docs.google.com/document/d/1MJ7R0YZRqa54vfT9ID8l5upATsqZ39I-82hATI4rCaQ/edit?tab=t.0",

        "🔄 Projets orchestrés & Alerting emaling":
            "https://docs.google.com/document/d/1YU93c_Gi0MvFghxcAAd7YXAnriuLxTK_C3lYrJmnhnY/edit?tab=t.0",

        "📌 Suivi demandes ponctuelles data (depuis 2026)":
            "https://www.notion.so/lapremierebrique/28c18ece1f2d80ab92cfc9b7fae8789c?v=28c18ece1f2d8033aa37000c8d2b563d",

        "🛠️ Guide de résolution de problèmes Power BI & Power Query":
            "https://docs.google.com/document/d/1nVkTPpOYRjmWPy_JOJHa3zw3l7maWpkU9FOkrY81jWY/edit?tab=t.0",

        "💡 Proposition de sujets data":
            "https://docs.google.com/document/d/1qeP5FdOSyZw0z_x9h4HEKaWP7OXxirL7SDt4cpgVJpg/edit?tab=t.0"
    }

    for label, url in DOCS.items():
        st.markdown(f"**{label}** : [{url}]({url})")
        
def page_emailing_courrier():
    st.title("📮 Emailing + Courrier PDP")

    st.markdown(
        """
        Application dédiée à la **gestion des PDP**, à l’envoi d’emails et de courriers, avec :
        - import d’un export Notion ZIP / CSV
        - recherche d’un **porteur de projet (PDP)** par email
        - affichage de la **fiche PDP complète** (identité, projet, contentieux, risque, hypothèque)
        - préparation et envoi d’**emails**
        - préparation et envoi de **courriers**
        - configuration **SMTP Gmail**
        - connexion **Merci Facteur**
        - suivi des envois

        👉 **Ouvrir l'application Streamlit :**  
        [https://lpb-emaling-courrier.streamlit.app/](https://lpb-emaling-courrier.streamlit.app/)
        """
    )
# =============================================================================
# 🧭 ROUTAGE PRINCIPAL
# =============================================================================

def main():
    with st.sidebar:
        st.markdown("## 🧱 Outils Data LPB")
        app_choice = st.radio(
            "Choix de l’outil",
            [
                "Data Hub (BO/Notion)",
                "Vérification des votes Airtable",
                "Suivi Invest TDF",
                "Découpe et recollement de fichiers",
                "Préqual + Emailing + Scoring ML",
                "Emailing + Courrier PDP",
                "Comparaison des rendements actifs",
                "Documentation",
            ],
            index=0,
        )
    #
    if app_choice == "Vérification des votes Airtable":
        page_votes()
    elif app_choice == "Suivi Invest TDF":
        page_suivi_invest_tdf()
    elif app_choice == "Découpe et recollement de fichiers":
        page_file_splitter()
    elif app_choice == "Préqual + Emailing + Scoring ML":
        page_prequalification()
    elif app_choice == "Emailing + Courrier PDP":
        page_emailing_courrier()
    elif app_choice == "Comparaison des rendements actifs":
        page_rendements()
    elif app_choice == "Documentation":
        page_docs()
    else:
        page_data_hub()

if __name__ == "__main__":
    main()

































































